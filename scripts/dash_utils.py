from __future__ import annotations

import csv
import hashlib
import json
import os
import pickle
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from contextlib import contextmanager
from pathlib import Path
from typing import Iterator
from typing import Optional

import numpy as np
from scipy.constants import h as _PLANCK_H

from qkd_sim.config.schema import WDMConfig
from qkd_sim.physical.signal import build_wdm_grid

_SCRIPT_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _SCRIPT_DIR.parent

# Diagnostic output gate. Enable via env var, e.g. PowerShell:
#   $env:DEBUG_MODE="true"; python scripts/plot_noise_dash_ch.py ...
_DEBUG_MODE: bool = os.environ.get("DEBUG_MODE", "").lower() in ("1", "true", "yes", "on")

# WDM 参数从 YAML 加载，保持唯一真值
_CONFIG_DIR = _PROJECT_ROOT / "src" / "qkd_sim" / "config"
_YAML_WDM_PATH = _CONFIG_DIR / "defaults" / "wdm_para" / "wdm_100ghz.yaml"
_YAML_FIBER_PATH = _CONFIG_DIR / "defaults" / "fiber_para" / "fiber_smf.yaml"


def _load_wdm_params() -> dict:
    from qkd_sim.config.schema import load_wdm_config
    cfg = load_wdm_config(_YAML_WDM_PATH)
    num_ch = cfg.num_channels
    if num_ch is None:
        num_ch = int(cfg.end_channel - cfg.start_channel + 1)
    return dict(
        start_freq=cfg.start_freq,
        start_channel=cfg.start_channel,
        end_channel=cfg.end_channel,
        channel_spacing=cfg.channel_spacing,
        B_s=cfg.B_s,
        B_q=cfg.B_q,
        data_rate_bps=cfg.data_rate_bps,
        P0=cfg.P0,
        beta_rolloff=cfg.beta_rolloff,
        ook_filter_order=cfg.ook_filter_order,
        num_channels=num_ch,
        # quantum_channel_indices 由 _build_wdm_config 单独传入，不加入此 dict
    )


# =============================================================================
# Classical channel strategy resolver
# =============================================================================

from dataclasses import dataclass, field


@dataclass
class ClassicalChannelConstraint:
    """A reserved region (e.g. sync channel, reference channel) with a protection bandwidth."""
    name: str           # identifier: "sync", "reference", ...
    channel: int        # 1-based ITU G.694.1 channel number (e.g. C33 = 33)
    bandwidth_ghz: float  # protection bandwidth [GHz]; no classical signal passband
                          # may overlap with (center - bw/2, center + bw/2)


@dataclass
class ClassicalChannelStrategy:
    """Classical channel placement strategy for plotting scripts."""
    name: str             # "equal_interval" | "interleave" | <future>
    reference_channel: int  # 1-based ITU G.694.1 channel number of reference channel (e.g. C35 = 35)
    num_classical: int   # desired number of classical channels
    reserved: list[ClassicalChannelConstraint] = field(default_factory=list)


def _passband_overlaps_reserved(ch_itn: float, res: ClassicalChannelConstraint) -> bool:
    """Check whether a classical channel's passband overlaps a reserved region.

    Classical channel passband: (ch_itn*100GHz - 50GHz, ch_itn*100GHz + 50GHz) GHz
    Reserved region:            (res.channel*100GHz - bw/2, res.channel*100GHz + bw/2) GHz
    """
    classic_left = ch_itn * 100e9 - 50e9
    classic_right = ch_itn * 100e9 + 50e9
    res_left = res.channel * 100e9 - res.bandwidth_ghz / 2 * 1e9
    res_right = res.channel * 100e9 + res.bandwidth_ghz / 2 * 1e9
    return classic_left < res_right and classic_right > res_left


def resolve_classical_indices(cfg: ClassicalChannelStrategy) -> list[float]:
    """Resolve classical channel ITU G.694.1 channel numbers from a placement strategy.

    Returns a list of 1-based ITU channel numbers (may be half-integers like 32.5 for interleave).
    Automatically excludes any candidate whose passband overlaps a reserved region.

    For equal_interval: searches greedily from ref-1 downward, skipping blocked channels,
    so a single blocked candidate does not cause an error — the next valid one is used instead.
    Raises ValueError only when the entire ITU range [1, ref-1] is exhausted before enough
    valid channels are found.

    For interleave: keeps the original fixed-offset generation; raises immediately if any
    generated candidate is out of [1, 61], and raises if not enough valid channels remain.
    """
    ref = float(cfg.reference_channel)  # ITU channel number (1-based)

    def _is_blocked(ch: float) -> bool:
        if any(ch == res.channel for res in cfg.reserved):
            return True
        if any(_passband_overlaps_reserved(ch, res) for res in cfg.reserved):
            return True
        return False

    if cfg.name == "equal_interval":
        # Greedy search: walk from ref-1 downward, skip blocked channels
        valid: list[float] = []
        ch = ref - 1.0
        while ch >= 1.0 and len(valid) < cfg.num_classical:
            if not _is_blocked(ch):
                valid.append(ch)
            ch -= 1.0
        if len(valid) < cfg.num_classical:
            raise ValueError(
                f"Strategy 'equal_interval' requested {cfg.num_classical} classical channels "
                f"but only {len(valid)} valid positions found in ITU range "
                f"[1, {int(ref) - 1}] after skipping reserved channels. Valid: {valid}"
            )
        return valid

    elif cfg.name == "interleave":
        # Fixed symmetric offsets; out-of-range is an immediate error
        n = cfg.num_classical
        candidates = [ref + (2 * i - (n - 1)) / 2.0 for i in range(n)]
        valid = []
        for ch in candidates:
            if _is_blocked(ch):
                continue
            if ch < 1.0 or ch > 61.0:
                raise ValueError(
                    f"Classical channel index {ch} (ITU C{int(ch)}) is outside "
                    f"the valid range [1, 61] (C01–C61)"
                )
            valid.append(ch)
        if len(valid) < cfg.num_classical:
            raise ValueError(
                f"Strategy 'interleave' requested {cfg.num_classical} classical channels "
                f"but only {len(valid)} valid positions remain after applying reserved "
                f"constraints. Candidates: {candidates}, Valid: {valid}"
            )
        return valid

    else:
        raise ValueError(f"Unknown classical channel strategy: {cfg.name!r}")


def _load_classical_channel_strategy(raw: dict) -> ClassicalChannelStrategy | None:
    """Parse classical_channel_strategy from YAML dict. Returns None if not present."""
    strat = raw.get("classical_channel_strategy")
    if strat is None:
        return None
    reserved = [
        ClassicalChannelConstraint(
            name=r["name"],
            channel=int(r["channel"]),
            bandwidth_ghz=float(r["bandwidth_ghz"]),
        )
        for r in strat.get("reserved", [])
    ]
    return ClassicalChannelStrategy(
        name=strat["name"],
        reference_channel=int(strat["reference_channel"]),
        num_classical=int(strat["num_classical"]),
        reserved=reserved,
    )


def _load_classical_indices() -> list[int]:
    """从 WDM YAML 读取 classical_channel_indices（1-based ITU 信道号）。

    classical_channel_indices 有值时直接返回；否则通过 classical_channel_strategy 策略生成。
    返回 1-based ITU 信道号列表，如 [39, 40, 41] 表示 C39/C40/C41。
    """
    import yaml
    with open(_YAML_WDM_PATH, encoding="utf-8") as f:
        raw = yaml.safe_load(f)

    raw_val = raw.get("classical_channel_indices")
    if raw_val is not None:
        # YAML 解析器将 [38, 39, 40] 解析为 [38.0, 39.0, 40.0]；
        # 统一转为 int（interleave 半整数保留 float）
        return [int(ch) if ch == int(ch) else ch for ch in raw_val]

    strat = _load_classical_channel_strategy(raw)
    if strat is not None:
        indices = resolve_classical_indices(strat)
        return [int(ch) if ch == int(ch) else ch for ch in indices]

    return []


def add_strategy_cli_args(parser):
    """给 argparse parser 添加经典信道策略参数。"""
    parser.add_argument(
        "--strategy-name",
        default=None,
        choices=["equal_interval", "interleave"],
        help="Classical channel placement strategy",
    )
    parser.add_argument(
        "--num-classical",
        type=int,
        default=None,
        help="Number of classical channels",
    )
    parser.add_argument(
        "--reference-channel",
        type=int,
        default=None,
        help="1-based ITU G.694.1 reference channel number (e.g. 35 = C35)",
    )
    parser.add_argument(
        "--skr-model",
        type=str,
        default=None,
        choices=["infinite", "approx_finite", "strict_finite"],
        help="SKR model for display (overrides default_skr_model from YAML)",
    )


def override_strategy_from_cli(strategy_name, num_classical, reference_channel):
    """用 CLI 参数覆盖 YAML 中的 classical_channel_strategy。

    reserved 列表保持 YAML 原始值，只覆盖 name/num_classical/reference_channel。
    前提：YAML 中必须已有 classical_channel_strategy 块（含 reserved 定义）。
    """
    import yaml

    with open(_YAML_WDM_PATH, encoding="utf-8") as f:
        raw = yaml.safe_load(f)

    strat = raw.get("classical_channel_strategy", {})
    if not strat:
        raise ValueError(
            "YAML 中没有 classical_channel_strategy 配置，无法使用策略 CLI 参数。"
            "请在 wdm_100ghz.yaml 中添加 classical_channel_strategy 块"
            "（含 reserved 定义），或使用 classical_channel_indices 配置。"
        )
    if strategy_name is not None:
        strat["name"] = strategy_name
    if num_classical is not None:
        strat["num_classical"] = num_classical
    if reference_channel is not None:
        strat["reference_channel"] = reference_channel

    cfg = _load_classical_channel_strategy({"classical_channel_strategy": strat})
    indices = resolve_classical_indices(cfg)
    return [int(ch) if ch == int(ch) else ch for ch in indices]


WDM_PARAMS = _load_wdm_params()
CLASSICAL_INDICES = _load_classical_indices()
NOISE_GRID_RESOLUTION_HZ = 1e8
ACTIVE_THRESHOLD_DB = -50.0  # FWM active frequency bin threshold [dB]
NOISE_FLOOR_W = 1e-23
FREQ_GRID_PADDING_FACTOR = 1.5

# --- SKR constants and model registry ---
_SKR_YAML_PATH = _CONFIG_DIR / "defaults" / "skr_para" / "bb84_config.yaml"
_SKR_MODEL_FNS: dict[str, tuple[callable, str]] = {}


def _init_skr_model_registry() -> dict[str, tuple[callable, str]]:
    """Lazy-init the SKR model function registry to avoid circular imports at module level."""
    global _SKR_MODEL_FNS
    if not _SKR_MODEL_FNS:
        from qkd_sim.physical.skr.skr_decoy_bb84 import (
            infinite_key_rate,
            approx_finite_key_rate,
            strict_finite_key_rate,
        )
        _SKR_MODEL_FNS = {
            "infinite": (
                lambda d, f, s, p, optimize_params=False: infinite_key_rate(d, f, s, p),
                "无限长密钥",
            ),
            "approx_finite": (
                lambda d, f, s, p, optimize_params=False: approx_finite_key_rate(d, f, s, p),
                "近似有限长",
            ),
            "strict_finite": (strict_finite_key_rate,  "严格有限长"),
        }
    return _SKR_MODEL_FNS

# --- Default SKR model key ---
# Loaded from model_comparison.yaml; fallback to "approx_finite" if not configured.
DEFAULT_SKR_MODEL_KEY: str = "approx_finite"

def _init_default_skr_model() -> str:
    """Load default SKR model from plot config YAML."""
    global DEFAULT_SKR_MODEL_KEY
    from qkd_sim.config.plot_config import load_default_skr_model
    DEFAULT_SKR_MODEL_KEY = load_default_skr_model()
    return DEFAULT_SKR_MODEL_KEY

_init_default_skr_model()

# 调制格式：由 Dash 脚本 --modulation 参数设置
MODULATION_FORMAT: str = "dp-16qam"
def _load_fiber_params() -> dict:
    from qkd_sim.config.schema import load_fiber_config
    cfg = load_fiber_config(_YAML_FIBER_PATH)
    return dict(
        alpha_dB_per_km=cfg.alpha_dB_per_km,
        gamma_per_W_km=cfg.gamma_per_W_km,
        D_ps_nm_km=cfg.D_ps_nm_km,
        D_slope_ps_nm2_km=cfg.D_slope_ps_nm2_km,
        L_km=cfg.L_km,
        A_eff=cfg.A_eff,
        rayleigh_coeff=cfg.rayleigh_coeff,
        T_kelvin=cfg.T_kelvin,
        length_km_samples=cfg.length_km_samples,
    )

_FIBER_CFG = _load_fiber_params()
FIBER_PARAMS = {k: v for k, v in _FIBER_CFG.items() if k != 'length_km_samples'}
LENGTHS_KM = np.array(_FIBER_CFG['length_km_samples'])
OSA_CSV_PATH = _PROJECT_ROOT / "data" / "osa"
_osa_rbw_hz: float | None = None
_osa_rbw_csv_path: Path | None = None


def _infer_rbw_from_csv(csv_path: Path) -> float:
    """Infer OSA RBW from CSV frequency step, rounded to 1 significant figure."""
    frequencies = []
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            frequencies.append(float(row["frequency_THz"]))
    f_steps = np.diff(frequencies) * 1e12  # THz -> Hz
    mean_step = float(np.mean(np.abs(f_steps)))
    if mean_step <= 0:
        raise ValueError(f"Invalid frequency step {mean_step} in {csv_path}")
    magnitude = 10 ** np.floor(np.log10(mean_step))
    rbw = magnitude * np.round(mean_step / magnitude)
    return float(rbw)


def _get_osa_rbw(csv_path: Path | None = None) -> float:
    """Get OSA RBW, inferring from CSV if not yet set.

    Parameters
    ----------
    csv_path : Path or None
        OSA CSV path to infer RBW from. If None, uses a default path.
    """
    global _osa_rbw_hz, _osa_rbw_csv_path
    if csv_path is not None and (_osa_rbw_hz is None or _osa_rbw_csv_path != csv_path):
        _osa_rbw_csv_path = csv_path
        _osa_rbw_hz = _infer_rbw_from_csv(csv_path)
    elif _osa_rbw_hz is None:
        # Fallback to default path for backward compatibility
        csv_path = OSA_CSV_PATH / "spectrum_OOK.csv"
        _osa_rbw_csv_path = csv_path
        _osa_rbw_hz = _infer_rbw_from_csv(csv_path)
    return _osa_rbw_hz


def _build_caption() -> str:
    """Build figure caption string from WDM_PARAMS and FIBER_PARAMS.

    Returns a multi-line string showing:
    - Channel spacing, data rate, OSA RBW
    - Classical channel positions
    - Fiber parameters (alpha, gamma, D)
    """
    ch_spacing_ghz = WDM_PARAMS["channel_spacing"] / 1e9
    data_rate_gbps = WDM_PARAMS["data_rate_bps"] / 1e9
    # CLASSICAL_INDICES 已是 1-based ITU 信道号（如 [39, 40, 41] = C39/C40/C41）
    classical_freqs = [
        WDM_PARAMS["start_freq"] + (itn - WDM_PARAMS["start_channel"]) * WDM_PARAMS["channel_spacing"]
        for itn in CLASSICAL_INDICES
    ]
    classical_labels = [f"C{itn}" for itn in CLASSICAL_INDICES]
    classical_desc = ", ".join(
        f"{lbl} ({freq / 1e12:.3f} THz)"
        for lbl, freq in zip(classical_labels, classical_freqs)
    )
    return (
        f"Channel spacing: {ch_spacing_ghz:.0f} GHz | "
        f"Data rate: {data_rate_gbps:.0f} Gbps | "
        f"OSA RBW: {_get_osa_rbw() / 1e9:.0f} GHz | "
        f"Classical channels: {classical_desc}"
    )


def _to_dbm(values_w: np.ndarray) -> np.ndarray:
    """Convert power [W] to dBm with NaN for non-positive values."""
    out = np.full_like(values_w, np.nan, dtype=np.float64)
    mask = values_w > 0
    out[mask] = 10.0 * np.log10(values_w[mask] / 1e-3)
    return out


# =============================================================================
# SKR utility functions
# =============================================================================

def load_skr_config_for_dash(profile: str = "custom"):
    """Load SKRConfig and FiberConfig for Dash SKR calculations."""
    from qkd_sim.config.schema import load_fiber_config, load_skr_config
    fiber_cfg = load_fiber_config(_YAML_FIBER_PATH)
    skr_cfg = load_skr_config(_SKR_YAML_PATH, profile=profile)
    return fiber_cfg, skr_cfg


def noise_power_to_p_noise(P_w: float, f_hz: float, R_rep: float) -> float:
    """Convert noise power [W] to per-pulse noise photon count probability.

    Uses Poisson model: p_noise = 1 - exp(-mu), where
    mu = P / (h * f * R_rep) is the average noise photons per pulse.
    """
    if P_w <= 0.0 or R_rep <= 0.0:
        return 0.0
    mu = P_w / (_PLANCK_H * f_hz * R_rep)
    return float(1.0 - np.exp(-mu))


def compute_skr_point(noise_w: float, distance_m: float, f_hz: float, fiber_cfg, skr_cfg,
                      optimize: bool = False, model_keys: list[str] | None = None) -> dict:
    """Compute SKR at a single (noise, distance, frequency) point.

    Parameters
    ----------
    model_keys : list[str] | None
        SKR model keys to compute. None = compute all three.
        Dash callers should pass [DEFAULT_SKR_MODEL_KEY] to avoid wasted
        strict_finite optimizer calls when only approx_finite is needed.

    Returns dict: {model_key: (skr_bps, skr_bit_per_pulse, qber)}
    """
    p_noise = noise_power_to_p_noise(noise_w, f_hz, skr_cfg.R_rep)
    skr_fns = _init_skr_model_registry()
    keys = model_keys if model_keys is not None else list(skr_fns.keys())
    results: dict = {}
    for mkey in keys:
        fn, _ = skr_fns[mkey]
        try:
            results[mkey] = fn(distance_m, fiber_cfg, skr_cfg, p_noise, optimize_params=optimize)
        except Exception:
            results[mkey] = (0.0, 0.0, float("nan"))
    return results


def compute_skr_vs_channel(
    sweep: dict,
    dist_m: float,
    quantum_center_freqs_hz: np.ndarray,
    fiber_cfg,
    skr_cfg,
    optimize: bool = False,
    model_keys: list[str] | None = None,
) -> dict:
    """Compute SKR vs quantum channel for a single fiber length.

    sweep: dict[model_key] = {fwd, bwd, x, x_kind, y_kind}
    quantum_center_freqs_hz: array of quantum channel center frequencies [Hz]
    model_keys: SKR models to compute (None = all three)

    Returns: dict[model_key][skr_model][direction] = (np.array(bps), np.array(bpp), np.array(qber))
    """
    skr_fns = _init_skr_model_registry()
    skr_keys = model_keys if model_keys is not None else list(skr_fns.keys())
    n_ch = len(quantum_center_freqs_hz)

    # Initialize result structure
    skr_result: dict = {}
    for model_key in sweep:
        skr_result[model_key] = {}
        for mkey in skr_keys:
            skr_result[model_key][mkey] = {
                "fwd": ([], [], []),  # bps, bpp, qber
                "bwd": ([], [], []),
            }

    for model_key, entry in sweep.items():
        # Prefer noise_only_* (pure FWM+SpRS, no classical signal) for SKR computation.
        # with_signal stores per-channel integrated values; other noise types fall back to fwd/bwd.
        noise_only_fwd = entry.get("noise_only_fwd")
        noise_only_bwd = entry.get("noise_only_bwd")

        if noise_only_fwd is not None and noise_only_bwd is not None and len(np.asarray(noise_only_fwd)) == n_ch:
            # Per-channel integrated noise power (B_q-bandwidth integration)
            fwd_interp = np.asarray(noise_only_fwd, dtype=np.float64)
            bwd_interp = np.asarray(noise_only_bwd, dtype=np.float64)
        else:
            # Fallback: interpolate from frequency grid or use channel-center point values
            fwd_w = np.asarray(noise_only_fwd if noise_only_fwd is not None else entry.get("fwd", []), dtype=np.float64)
            bwd_w = np.asarray(noise_only_bwd if noise_only_bwd is not None else entry.get("bwd", []), dtype=np.float64)
            x_data = np.asarray(entry.get("x", []), dtype=np.float64)
            x_kind = entry.get("x_kind", "")

            if len(x_data) == 0 or len(fwd_w) == 0:
                continue

            if x_kind == "frequency_grid" and len(x_data) > 1:
                fwd_interp = np.interp(quantum_center_freqs_hz, x_data, fwd_w)
                bwd_interp = np.interp(quantum_center_freqs_hz, x_data, bwd_w)
            elif len(x_data) == n_ch:
                fwd_interp = fwd_w.copy()
                bwd_interp = bwd_w.copy()
            else:
                continue

        for direction, noise_arr in [("fwd", fwd_interp), ("bwd", bwd_interp)]:
            for ch_i in range(n_ch):
                skr_d = compute_skr_point(
                    float(noise_arr[ch_i]), dist_m,
                    float(quantum_center_freqs_hz[ch_i]), fiber_cfg, skr_cfg, optimize,
                    model_keys=skr_keys,
                )
                for mkey in skr_keys:
                    bps, bpp, qber = skr_d.get(mkey, (0.0, 0.0, float("nan")))
                    lists = skr_result[model_key][mkey][direction]
                    lists[0].append(bps)
                    lists[1].append(bpp)
                    lists[2].append(qber)

    # Convert lists to arrays
    for model_key in skr_result:
        for mkey in skr_result[model_key]:
            for direction in ("fwd", "bwd"):
                lists = skr_result[model_key][mkey][direction]
                skr_result[model_key][mkey][direction] = (
                    np.array(lists[0]), np.array(lists[1]), np.array(lists[2]),
                )

    return skr_result


def compute_skr_cache_for_power(
    power_dbm: float,
    sweep_at_l: dict,
    l_idx: int,
    length_km: float,
    quantum_center_freqs_hz: np.ndarray,
    fiber_cfg,
    skr_cfg,
    model_keys: list[str] | None = None,
) -> dict:
    """Build SKR cache for a single (power, length) combination.

    Fully parameterized, no closures — suitable for ProcessPoolExecutor
    on all platforms. Thin wrapper around ``compute_skr_vs_channel``.

    Parameters
    ----------
    power_dbm : float
        Classical channel power (for logging / future use).
    sweep_at_l : dict
        Precomputed noise sweep at one (power, length) combo.
    l_idx : int
        Length index (for caller bookkeeping).
    length_km : float
        Fiber length [km].
    quantum_center_freqs_hz : ndarray
        Quantum channel center frequencies [Hz].
    fiber_cfg : FiberConfig
    skr_cfg : SKRConfig
    model_keys : list[str] | None
        SKR model keys to compute (None = all three).

    Returns
    -------
    dict — same format as ``compute_skr_vs_channel``.
    """
    dist_m = float(length_km * 1000.0)
    return compute_skr_vs_channel(
        sweep_at_l, dist_m, quantum_center_freqs_hz,
        fiber_cfg, skr_cfg, optimize=skr_cfg.optimize_params,
        model_keys=model_keys,
    )


def compute_skr_vs_length(
    sweep: dict,
    ch_idx: int,
    lengths_km: np.ndarray,
    fiber_cfg,
    skr_cfg,
    optimize: bool = False,
    model_keys: list[str] | None = None,
) -> dict:
    """Compute SKR vs fiber length for a single quantum channel.

    sweep: dict[model_key] = {fwd, bwd, x, x_kind, y_kind} where y is per-length
    ch_idx: ITU G.694.1 channel index
    model_keys: SKR models to compute (None = all three)

    Returns: dict[model_key][skr_model][direction] = (np.array(bps), np.array(bpp), np.array(qber)) per length
    """
    skr_fns = _init_skr_model_registry()
    skr_keys = model_keys if model_keys is not None else list(skr_fns.keys())
    skr_result: dict = {}

    # Get channel center frequency
    f_q_hz = WDM_PARAMS["start_freq"] + (ch_idx - WDM_PARAMS["start_channel"]) * WDM_PARAMS["channel_spacing"]

    for model_key, entry in sweep.items():
        # Prefer noise_only_* for SKR; with_signal stores these to exclude classical signal power.
        fwd_w = np.asarray(entry.get("noise_only_fwd", entry.get("fwd", [])), dtype=np.float64)
        bwd_w = np.asarray(entry.get("noise_only_bwd", entry.get("bwd", [])), dtype=np.float64)

        n_len = min(len(fwd_w), len(lengths_km))
        if n_len == 0:
            continue

        skr_by_model: dict = {}
        for mkey in skr_keys:
            fwd_bps, fwd_bpp, fwd_qber = [], [], []
            bwd_bps, bwd_bpp, bwd_qber = [], [], []
            for li in range(n_len):
                fwd_d = compute_skr_point(float(fwd_w[li]), float(lengths_km[li] * 1000), f_q_hz, fiber_cfg, skr_cfg, optimize, model_keys=skr_keys)
                bwd_d = compute_skr_point(float(bwd_w[li]), float(lengths_km[li] * 1000), f_q_hz, fiber_cfg, skr_cfg, optimize, model_keys=skr_keys)
                bps_f, bpp_f, q_f = fwd_d.get(mkey, (0.0, 0.0, float("nan")))
                bps_b, bpp_b, q_b = bwd_d.get(mkey, (0.0, 0.0, float("nan")))
                fwd_bps.append(bps_f); fwd_bpp.append(bpp_f); fwd_qber.append(q_f)
                bwd_bps.append(bps_b); bwd_bpp.append(bpp_b); bwd_qber.append(q_b)
            skr_by_model[mkey] = {
                "fwd": (np.array(fwd_bps), np.array(fwd_bpp), np.array(fwd_qber)),
                "bwd": (np.array(bwd_bps), np.array(bwd_bpp), np.array(bwd_qber)),
            }
        skr_result[model_key] = skr_by_model

    return skr_result

FIBER_PARAMS = dict(
    alpha_dB_per_km=0.2,
    gamma_per_W_km=1.3,
    D_ps_nm_km=17.0,
    D_slope_ps_nm2_km=0.056,
    L_km=50.0,
    A_eff=80e-12,
    rayleigh_coeff=4.8e-8,
    T_kelvin=300.0,
)

# --- Multiprocessing for precomputation ---
# Disabled on Windows (nt): scipy Cython DLLs fail to load in spawned subprocesses.
# Enable via QKD_DASH_MP=1 env var on non-Windows systems for MP speedup.
_MP_ENABLED = os.environ.get("QKD_DASH_MP", "1") != "0" and os.name != "nt"
_PROFILE_ENABLED = os.environ.get("QKD_DASH_PROFILE", "1").lower() not in ("0", "false", "no")
_CSV_CACHE_ENABLED = os.environ.get("QKD_DASH_CSV_CACHE", "1").lower() not in ("0", "false", "no")
_PROFILE_INDENT = 0
_MODEL_KEY_FILTER: list[str] | None = None


def set_model_key_filter(model_keys: list[str] | None) -> None:
    """Limit Dash precomputation to a subset of configured model keys."""
    global _MODEL_KEY_FILTER
    _MODEL_KEY_FILTER = list(model_keys) if model_keys else None


def _apply_model_key_filter(model_keys: list[str]) -> list[str]:
    if _MODEL_KEY_FILTER is None:
        return model_keys
    requested = set(_MODEL_KEY_FILTER)
    selected = [mk for mk in model_keys if mk in requested]
    unknown = sorted(requested - set(model_keys))
    if unknown:
        raise ValueError(
            f"Unknown model key(s) for modulation {MODULATION_FORMAT!r}: {unknown}. "
            f"Available: {model_keys}"
        )
    if not selected:
        raise ValueError("Model filter selected no models")
    return selected


def describe_compute_device() -> str:
    """Return the active compute backend used by Dash noise precomputation."""
    try:
        from qkd_sim.utils.gpu_utils import get_array_module, has_cupy

        if has_cupy():
            xp = get_array_module()
            device_id = xp.cuda.runtime.getDevice()
            props = xp.cuda.runtime.getDeviceProperties(device_id)
            name = props.get("name", b"CUDA GPU")
            if isinstance(name, bytes):
                name = name.decode("utf-8", errors="replace")
            return f"CUDA GPU: {name} (CuPy; FWM continuous spectrum path)"
    except Exception as exc:
        return f"CPU (NumPy/SciPy; GPU detection failed: {exc})"
    return "CPU (NumPy/SciPy)"


def print_compute_device() -> None:
    """Print compute backend once at script startup."""
    print(f"Compute device: {describe_compute_device()}")


def ensure_port_free(port: int, host: str = "127.0.0.1") -> None:
    """Exit with an actionable error if ``host:port`` is already bound.

    Guards against a stale Dash process silently hijacking a new launch —
    if the port is held, ``app.run`` would emit an OSError that is easy to
    miss, leaving the browser connected to the zombie.
    """
    import socket
    import subprocess
    import sys

    probe = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        probe.bind((host, port))
    except OSError:
        pid_hint = ""
        try:
            result = subprocess.run(
                ["netstat", "-ano"], capture_output=True, text=True, timeout=5
            )
            for line in result.stdout.splitlines():
                parts = line.split()
                if (
                    len(parts) >= 5
                    and parts[0] == "TCP"
                    and parts[1].endswith(f":{port}")
                    and parts[3] == "LISTENING"
                ):
                    pid_hint = f" (held by PID {parts[4]})"
                    break
        except Exception:
            pass
        sys.stderr.write(
            f"\n[ERROR] Port {port} is already in use{pid_hint}.\n"
            f"        A previous Dash instance is still running and will keep\n"
            f"        serving stale --type / --modulation results to the browser.\n"
            f"        Kill it first, e.g. (PowerShell):\n"
            f"          Stop-Process -Id <PID> -Force\n"
            f"        Then re-run this script.\n\n"
        )
        sys.exit(1)
    finally:
        probe.close()


def set_csv_cache_enabled(enabled: bool) -> None:
    """Enable or disable Dash precompute CSV cache I/O."""
    global _CSV_CACHE_ENABLED
    _CSV_CACHE_ENABLED = bool(enabled)


@contextmanager
def profile_scope(label: str) -> Iterator[None]:
    """Print elapsed time for a named Dash precomputation step."""
    global _PROFILE_INDENT
    if not _PROFILE_ENABLED:
        yield
        return

    indent = "  " * _PROFILE_INDENT
    _PROFILE_INDENT += 1
    t_start = time.perf_counter()
    try:
        yield
    finally:
        elapsed = time.perf_counter() - t_start
        _PROFILE_INDENT -= 1
        print(f"[profile] {indent}{label}: {elapsed:.3f}s")


def _get_mp_workers() -> int:
    """Number of worker processes for precomputation."""
    return max(1, os.cpu_count() or 1)


_GPU_STATUS_PRINTED: bool = False


def _print_gpu_status_once() -> None:
    """One-shot diagnostic print of GPU availability and free VRAM (DEBUG_MODE only)."""
    global _GPU_STATUS_PRINTED
    if _GPU_STATUS_PRINTED:
        return
    _GPU_STATUS_PRINTED = True
    if not _DEBUG_MODE:
        return
    try:
        from qkd_sim.utils.gpu_utils import GPU_ENABLED
        if GPU_ENABLED:
            try:
                import cupy as cp
                free_b, total_b = cp.cuda.Device().mem_info
                dev_id = cp.cuda.Device().id
                print(
                    f"[gpu] is_gpu=True device={dev_id} "
                    f"free={free_b / 1024**3:.2f} GiB / total={total_b / 1024**3:.2f} GiB"
                )
            except Exception as exc:
                print(f"[gpu] is_gpu=True (mem_info unavailable: {exc!r})")
        else:
            print("[gpu] is_gpu=False — solvers will fall back to CPU NumPy path")
    except Exception as exc:
        print(f"[gpu] status check failed: {exc!r}")


# --- Worker functions for multiprocessing (must be module-level for Windows spawn) ---

def _precompute_length_worker(
    power_dbm: float,
    noise_type: str,
    specs: dict,
    LENGTHS_KM: np.ndarray,
    base_config: WDMConfig,
    noise_f_grid: np.ndarray,
    osa_csv_path: Path,
    fiber_params: dict,
    osa_center_freq_hz: float | None = None,
) -> tuple[float, dict, list]:
    """Worker: precompute by length for one power level. Runs in subprocess."""
    set_power_override(float(power_dbm))
    all_by_idx, valid_ch = precompute_by_length(
        noise_type, specs, LENGTHS_KM, base_config,
        noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
    )
    return float(power_dbm), all_by_idx, valid_ch


def _precompute_channel_worker(
    power_dbm: float,
    noise_type: str,
    specs: dict,
    LENGTHS_KM: np.ndarray,
    base_config: WDMConfig,
    noise_f_grid: np.ndarray,
    osa_csv_path: Path,
    fiber_params: dict,
    osa_center_freq_hz: float | None = None,
) -> tuple[float, dict, list]:
    """Worker: precompute by channel for one power level. Runs in subprocess."""
    set_power_override(float(power_dbm))
    all_by_idx, valid_l = precompute_by_channel(
        noise_type, specs, LENGTHS_KM, base_config,
        noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
    )
    return float(power_dbm), all_by_idx, valid_l


# --- Power levels for startup precomputation (step=1 dBm, -15 to 0 dBm) ---
PRECOMPUTE_POWER_LEVELS = [float(p) for p in range(-15, 1)]


# --- Clear stale CSV cache ---
def _clear_precomputed_csv_files(index_prefix: str | None = None) -> None:
    """Delete CSV files in precomputed dir to avoid stale data from changed params.

    index_prefix: if None, delete all; if "ch" or "len", only that subset.
    """
    if not _CSV_CACHE_ENABLED:
        return
    skipped: list[str] = []
    for f in _PRECOMPUTED_DIR.glob("*.csv"):
        if index_prefix is None or f.name.startswith(index_prefix):
            try:
                f.unlink()
            except PermissionError:
                skipped.append(f.name)
    if skipped:
        preview = ", ".join(skipped[:3])
        suffix = "" if len(skipped) <= 3 else f", ... +{len(skipped) - 3} more"
        print(f"[profile] warning: skipped {len(skipped)} locked CSV cache files: {preview}{suffix}")


# --- Power override for classical channel launch power ---
_POWER_OVERRIDE_DBM: Optional[float] = None


def set_power_override(dbm: Optional[float]) -> None:
    """Override P0 (classical channel launch power in dBm).

    Set to None to revert to YAML-configured default.
    """
    global _POWER_OVERRIDE_DBM
    _POWER_OVERRIDE_DBM = dbm


def _get_P0() -> float:
    """Return effective P0 in linear Watts."""
    if _POWER_OVERRIDE_DBM is not None:
        return 1e-3 * 10 ** (_POWER_OVERRIDE_DBM / 10.0)
    return WDM_PARAMS["P0"]


# --- Power-based CSV caching for instant slider response ---
_PRECOMPUTED_DIR = _PROJECT_ROOT / "data" / "precomputed"
_POWER_CACHE: dict[float, dict] = {}  # power_dbm -> precomputed result dict


def _power_csv_path(noise_type: str, model_key: str, direction: str, power_dbm: float, index_prefix: str = "ch") -> Path:
    """CSV path for a specific (noise_type, model, direction, power, index_prefix) combination."""
    return _PRECOMPUTED_DIR / f"{index_prefix}_{noise_type}_{model_key}_{direction}_p{int(power_dbm):+d}.csv"


def _save_power_csv(data: np.ndarray, filepath: Path) -> None:
    """Save 2D power array to CSV (rows = outer index, cols = inner index)."""
    import csv

    if not _CSV_CACHE_ENABLED:
        return
    _PRECOMPUTED_DIR.mkdir(parents=True, exist_ok=True)
    try:
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in data:
                writer.writerow([f"{v:.6e}" for v in row])
    except PermissionError as exc:
        print(f"[profile] warning: skip CSV cache write for {filepath.name}: {exc}")


def _load_power_csv(filepath: Path) -> np.ndarray:
    """Load 2D power array from CSV."""
    import csv

    rows = []
    with open(filepath, encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append([float(x) for x in row])
    return np.array(rows, dtype=np.float64)


def _cache_precomputed_result(
    all_by_idx: dict, power_dbm: float, noise_type: str, model_keys: list[str], index_prefix: str = "ch"
) -> None:
    """Save precomputed result to CSV for each (model, direction) at given power.

    index_prefix: "ch" for precompute_by_channel (rows = length indices), "len" for precompute_by_length (rows = channel indices).
    """
    if not all_by_idx:
        return
    first_key = min(all_by_idx.keys())
    # Only process model keys actually present in the data (model_keys may
    # include entries filtered out by --models, etc.)
    available = [mk for mk in model_keys if mk in all_by_idx[first_key]]

    for mk in available:
        ref_shape = all_by_idx[first_key][mk]["fwd"].shape
        if not all(all_by_idx[k][mk]["fwd"].shape == ref_shape for k in all_by_idx):
            continue  # Skip mixed-shape models
        for direction in ("fwd", "bwd"):
            data = np.array(
                [all_by_idx[k][mk][direction] for k in sorted(all_by_idx.keys())],
                dtype=np.float64,
            )
            _save_power_csv(data, _power_csv_path(noise_type, mk, direction, power_dbm, index_prefix))


def _load_cached_power(noise_type: str, model_keys: list[str], index_prefix: str = "ch") -> tuple[dict | None, float]:
    """Try to load cached precomputed result for current power override.

    index_prefix: "ch" for length-indexed data (precompute_by_channel), "len" for channel-indexed data (precompute_by_length).
    """
    power = _POWER_OVERRIDE_DBM if _POWER_OVERRIDE_DBM is not None else 0.0
    # Try exact power match in memory cache (distinguish by prefix)
    cache_key = (index_prefix, power)
    if cache_key in _POWER_CACHE:
        return _POWER_CACHE[cache_key], power
    # Try CSV files for this power (only for continuous models)
    # Find first continuous model as reference
    first_cont_mk = None
    for mk in model_keys:
        fwd_path = _power_csv_path(noise_type, mk, "fwd", power, index_prefix)
        if fwd_path.exists():
            first_cont_mk = mk
            break
    if first_cont_mk is None:
        return None, power  # No cached continuous models
    # Load continuous models from CSV
    # For "ch": outer index = length (n_l=18), data shape = (n_l, n_f)
    # For "len": outer index = channel (n_ch=61), data shape = (n_ch, n_l)
    # Determine outer dimension from the first CSV's row count
    ref_path = _power_csv_path(noise_type, first_cont_mk, "fwd", power, index_prefix)
    ref_data = _load_power_csv(ref_path)
    n_outer = ref_data.shape[0]

    all_by_idx = {}
    for idx in range(n_outer):
        all_by_idx[idx] = {}
        for mk in model_keys:
            fwd_path = _power_csv_path(noise_type, mk, "fwd", power, index_prefix)
            bwd_path = _power_csv_path(noise_type, mk, "bwd", power, index_prefix)
            if fwd_path.exists() and bwd_path.exists():
                fwd_data = _load_power_csv(fwd_path)
                bwd_data = _load_power_csv(bwd_path)
                all_by_idx[idx][mk] = {
                    "fwd": fwd_data[idx] if idx < len(fwd_data) else np.array([]),
                    "bwd": bwd_data[idx] if idx < len(bwd_data) else np.array([]),
                    "x": np.array([]),
                    "x_kind": "",
                    "y_kind": "",
                }
    _POWER_CACHE[cache_key] = all_by_idx
    return all_by_idx, power


def _persistent_cache_path(cache_name: str, payload: dict) -> Path:
    """Return a parameter-hash cache path for expensive Dash precompute results."""
    encoded = json.dumps(payload, sort_keys=True, default=str).encode("utf-8")
    digest = hashlib.sha256(encoded).hexdigest()[:16]
    return _PRECOMPUTED_DIR / f"{cache_name}_{digest}.pkl"


def _load_persistent_cache(cache_name: str, payload: dict) -> dict | None:
    if not _CSV_CACHE_ENABLED:
        return None
    path = _persistent_cache_path(cache_name, payload)
    if not path.exists():
        return None
    try:
        with open(path, "rb") as f:
            cached = pickle.load(f)
    except Exception as exc:
        print(f"[profile] warning: ignore unreadable persistent cache {path.name}: {exc}")
        return None
    print(f"[profile] persistent cache hit: {path.name}")
    return cached


def _save_persistent_cache(cache_name: str, payload: dict, data: dict) -> None:
    if not _CSV_CACHE_ENABLED:
        return
    _PRECOMPUTED_DIR.mkdir(parents=True, exist_ok=True)
    path = _persistent_cache_path(cache_name, payload)
    try:
        with open(path, "wb") as f:
            pickle.dump(data, f, protocol=pickle.HIGHEST_PROTOCOL)
    except PermissionError as exc:
        print(f"[profile] warning: skip persistent cache write for {path.name}: {exc}")


def _with_signal_cache_payload(
    noise_type: str,
    model_keys: list[str],
    LENGTHS_KM: np.ndarray,
    base_config: WDMConfig,
    noise_f_grid: np.ndarray,
    osa_csv_path: Path,
    fiber_params: dict,
    osa_center_freq_hz: float | None,
) -> dict:
    osa_stat = osa_csv_path.stat() if osa_csv_path.exists() else None
    return {
        "schema": 1,
        "noise_type": noise_type,
        "modulation": MODULATION_FORMAT,
        "model_keys": list(model_keys),
        "power_levels": [float(p) for p in PRECOMPUTE_POWER_LEVELS],
        "lengths_km": np.asarray(LENGTHS_KM, dtype=np.float64).tolist(),
        "noise_grid": {
            "size": int(noise_f_grid.size),
            "first": float(noise_f_grid[0]) if noise_f_grid.size else None,
            "last": float(noise_f_grid[-1]) if noise_f_grid.size else None,
            "step": float(np.mean(np.diff(noise_f_grid))) if noise_f_grid.size > 1 else None,
        },
        "active_threshold_db": float(ACTIVE_THRESHOLD_DB),
        "classical_indices": list(CLASSICAL_INDICES),
        "wdm": {
            "start_freq": float(base_config.start_freq),
            "start_channel": float(base_config.start_channel),
            "end_channel": float(base_config.end_channel),
            "channel_spacing": float(base_config.channel_spacing),
            "B_s": float(base_config.B_s),
            "B_q": float(base_config.B_q),
            "data_rate_bps": float(base_config.data_rate_bps),
            "P0": float(_get_P0()),
            "beta_rolloff": float(base_config.beta_rolloff),
            "ook_filter_order": int(base_config.ook_filter_order),
            "quantum_channel_indices": list(base_config.quantum_channel_indices),
            "num_channels": int(base_config.num_channels),
        },
        "fiber": {k: float(v) for k, v in fiber_params.items()},
        "osa": {
            "path": str(osa_csv_path),
            "mtime_ns": int(osa_stat.st_mtime_ns) if osa_stat else None,
            "size": int(osa_stat.st_size) if osa_stat else None,
            "center_freq_hz": float(osa_center_freq_hz) if osa_center_freq_hz is not None else None,
        },
    }


def _scale_precomputed_result(all_by_idx: dict, scale: float) -> dict:
    """Scale cached fwd/bwd noise arrays while preserving x-axis metadata."""
    scaled: dict = {}
    for outer_idx, model_data in all_by_idx.items():
        scaled[outer_idx] = {}
        for model_key, entry in model_data.items():
            scaled[outer_idx][model_key] = {
                "fwd": np.asarray(entry.get("fwd", []), dtype=np.float64) * scale,
                "bwd": np.asarray(entry.get("bwd", []), dtype=np.float64) * scale,
                "x": np.asarray(entry.get("x", []), dtype=np.float64),
                "x_kind": entry.get("x_kind", ""),
                "y_kind": entry.get("y_kind", ""),
            }
    return scaled


def _combine_precomputed_results(
    fwm_by_idx: dict,
    sprs_by_idx: dict,
    fwm_scale: float,
    sprs_scale: float,
) -> dict:
    """Combine separately precomputed FWM and SpRS data for ``both`` plots."""
    combined: dict = {}
    for outer_idx in sorted(set(fwm_by_idx.keys()) & set(sprs_by_idx.keys())):
        combined[outer_idx] = {}
        model_keys = sorted(set(fwm_by_idx[outer_idx].keys()) & set(sprs_by_idx[outer_idx].keys()))
        for model_key in model_keys:
            fwm_entry = fwm_by_idx[outer_idx][model_key]
            sprs_entry = sprs_by_idx[outer_idx][model_key]
            combined[outer_idx][model_key] = {
                "fwd": (
                    np.asarray(fwm_entry.get("fwd", []), dtype=np.float64) * fwm_scale
                    + np.asarray(sprs_entry.get("fwd", []), dtype=np.float64) * sprs_scale
                ),
                "bwd": (
                    np.asarray(fwm_entry.get("bwd", []), dtype=np.float64) * fwm_scale
                    + np.asarray(sprs_entry.get("bwd", []), dtype=np.float64) * sprs_scale
                ),
                "x": np.asarray(fwm_entry.get("x", sprs_entry.get("x", [])), dtype=np.float64),
                "x_kind": fwm_entry.get("x_kind", sprs_entry.get("x_kind", "")),
                "y_kind": fwm_entry.get("y_kind", sprs_entry.get("y_kind", "")),
            }
    return combined


def _power_scaling_factor(noise_type: str, power_dbm: float) -> float | None:
    """Return exact pump-power scaling from the 0 dBm precompute."""
    if noise_type == "sprs":
        return float(10.0 ** (power_dbm / 10.0))
    if noise_type == "fwm":
        return float(10.0 ** (3.0 * power_dbm / 10.0))
    return None


_LEGEND_SYNC_JS = """
(function() {
    function getLegendGroup(el, curveNumber) {
        var fd = el._fullData || [];
        return (fd[curveNumber] && fd[curveNumber].legendgroup) ? fd[curveNumber].legendgroup : null;
    }

    function getLegendGroups(el) {
        var groups = [];
        var seen = {};
        var data = el.data || [];
        for (var i = 0; i < data.length; i++) {
            var grp = data[i].legendgroup;
            if (!grp || seen[grp]) { continue; }
            seen[grp] = true;
            groups.push(grp);
        }
        return groups;
    }

    function setAllGroupsVisible(el) {
        var data = el.data || [];
        for (var i = 0; i < data.length; i++) {
            if (data[i].legendgroup) {
                data[i].visible = true;
            }
        }
    }

    function isolateGroup(el, targetGroup) {
        var data = el.data || [];
        for (var i = 0; i < data.length; i++) {
            var grp = data[i].legendgroup;
            if (!grp) { continue; }
            data[i].visible = (grp === targetGroup) ? true : 'legendonly';
        }
    }

    function isGroupIsolated(el, targetGroup) {
        var groups = getLegendGroups(el);
        if (groups.length <= 1) { return false; }

        var visibleGroups = {};
        var data = el.data || [];
        for (var i = 0; i < data.length; i++) {
            var grp = data[i].legendgroup;
            if (!grp || data[i].visible === 'legendonly') { continue; }
            visibleGroups[grp] = true;
        }

        var visibleCount = 0;
        for (var j = 0; j < groups.length; j++) {
            if (visibleGroups[groups[j]]) {
                visibleCount += 1;
            }
        }

        return visibleCount === 1 && !!visibleGroups[targetGroup];
    }

    function attachLegendSync() {
        var el = document.querySelector('.js-plotly-plot');
        if (!el) { setTimeout(attachLegendSync, 500); return; }
        if (el.dataset.legendSyncAttached === '1') { return; }
        el.dataset.legendSyncAttached = '1';

        el.on('plotly_legendclick', function(ev) {
            var grp = getLegendGroup(el, ev.curveNumber);
            if (!grp) { return true; }

            var on = false;
            var data = el.data || [];
            for (var i = 0; i < data.length; i++) {
                if (data[i].legendgroup === grp && data[i].visible !== 'legendonly') {
                    on = true;
                    break;
                }
            }

            var nv = on ? 'legendonly' : true;
            for (var j = 0; j < data.length; j++) {
                if (data[j].legendgroup === grp) {
                    data[j].visible = nv;
                }
            }
            Plotly.redraw(el);
            return false;
        });

        el.on('plotly_legenddoubleclick', function(ev) {
            var grp = getLegendGroup(el, ev.curveNumber);
            if (!grp) { return true; }

            var groups = getLegendGroups(el);
            if (groups.length <= 1) { return false; }

            if (isGroupIsolated(el, grp)) {
                setAllGroupsVisible(el);
            } else {
                isolateGroup(el, grp);
            }

            Plotly.redraw(el);
            return false;
        });
    }

    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', attachLegendSync);
    } else { attachLegendSync(); }
})();
"""


def _resolve_osa_csv(modulation_format: str = "DP-16QAM") -> tuple[Path, float | None]:
    """Resolve OSA CSV path based on modulation format, and extract center frequency from filename.

    Filename pattern: spectrum_<fmt>_<freq>THz.csv (e.g., spectrum_16QAM_196.0THz.csv)

    Parameters
    ----------
    modulation_format : str
        "OOK" or "DP-16QAM" (case-insensitive)

    Returns
    -------
    csv_path : Path
        Path to the OSA CSV file
    center_freq_hz : float | None
        Center frequency in Hz parsed from filename, or None if not found
    """
    import re

    fmt_lower = modulation_format.lower()
    # Map modulation format to filename pattern suffix
    if fmt_lower == "ook":
        pattern = re.compile(r"^spectrum_OOK_(\d+\.\d+)THz\.csv$", re.IGNORECASE)
    else:  # dp-16qam or default
        pattern = re.compile(r"^spectrum_16QAM_(\d+\.\d+)THz\.csv$", re.IGNORECASE)

    for csv_path in OSA_CSV_PATH.glob("*.csv"):
        m = pattern.match(csv_path.name)
        if m:
            center_freq_thz = float(m.group(1))
            center_freq_hz = center_freq_thz * 1e12
            return csv_path, center_freq_hz

    # Fallback: try old filename without frequency suffix
    if fmt_lower == "ook":
        filename = "spectrum_OOK.csv"
    else:
        filename = "spectrum_16QAM.csv"
    csv_path = OSA_CSV_PATH / filename
    if csv_path.exists():
        return csv_path, None  # Old file, no center frequency
    raise FileNotFoundError(f"OSA CSV not found for {modulation_format!r} in {OSA_CSV_PATH}")


def _build_noise_frequency_grid(config: WDMConfig) -> np.ndarray:
    half_span = (config.num_channels - 1) / 2.0 * config.channel_spacing
    center_freq = config.start_freq + half_span
    padding = FREQ_GRID_PADDING_FACTOR * config.channel_spacing
    f_min = center_freq - half_span - padding
    f_max = center_freq + half_span + padding
    n_points = int(np.ceil((f_max - f_min) / NOISE_GRID_RESOLUTION_HZ)) + 1
    return np.linspace(f_min, f_max, n_points)


def _build_wdm_config(quantum_indices: list[int]) -> WDMConfig:
    params = {k: v for k, v in WDM_PARAMS.items() if k != "P0"}
    return WDMConfig(**params, quantum_channel_indices=list(quantum_indices), P0=_get_P0())


def _modulation_format_for_model(model_key: str, spec: dict) -> str:
    """Determine modulation_format parameter for build_wdm_grid based on model and global setting."""
    from qkd_sim.physical.signal import SpectrumType
    if spec["spectrum_type"] == SpectrumType.SINGLE_FREQ:
        return "OOK"  # discrete model — no spectral shape dependency
    if spec["spectrum_type"] == SpectrumType.NRZ_OOK:
        return "OOK"
    if spec["spectrum_type"] == SpectrumType.RAISED_COSINE:
        return "DP-16QAM"
    # OSA_SAMPLED: determine by model_key
    if model_key == "osa_ook":
        return "OOK"
    return "DP-16QAM"


def _build_model_grid(
    model_key: str,
    spec: dict,
    base_config: WDMConfig,
    f_grid: np.ndarray,
    osa_csv_path: Path,
    osa_center_freq_hz: float | None = None,
):
    from qkd_sim.physical.signal import SpectrumType

    # Respect base_config.quantum_channel_indices directly when it is a single
    # channel (from _build_wdm_config([q_idx]) in the discrete model's
    # per-channel grid loop — preserves per-channel isolation).
    # For multi-channel configs (e.g. continuous model), recompute from
    # current CLASSICAL_INDICES so that CLI overrides take effect.
    if len(base_config.quantum_channel_indices) == 1:
        current_quantum_indices = list(base_config.quantum_channel_indices)
    else:
        all_itn = list(range(int(base_config.start_channel), int(base_config.end_channel) + 1))
        current_quantum_indices = sorted(ch for ch in all_itn if ch not in CLASSICAL_INDICES)

    if spec["beta_rolloff"] is not None:
        model_config = WDMConfig(
            start_freq=base_config.start_freq,
            start_channel=base_config.start_channel,
            end_channel=base_config.end_channel,
            channel_spacing=base_config.channel_spacing,
            B_s=base_config.B_s,
            data_rate_bps=base_config.data_rate_bps,
            P0=_get_P0(),
            beta_rolloff=spec["beta_rolloff"],
            ook_filter_order=base_config.ook_filter_order,
            quantum_channel_indices=current_quantum_indices,
            channel_powers_W=base_config.channel_powers_W,
            num_channels=int(base_config.num_channels),
        )
    else:
        model_config = WDMConfig(
            start_freq=base_config.start_freq,
            start_channel=base_config.start_channel,
            end_channel=base_config.end_channel,
            channel_spacing=base_config.channel_spacing,
            B_s=base_config.B_s,
            data_rate_bps=base_config.data_rate_bps,
            P0=_get_P0(),
            beta_rolloff=base_config.beta_rolloff,
            ook_filter_order=base_config.ook_filter_order,
            quantum_channel_indices=current_quantum_indices,
            channel_powers_W=base_config.channel_powers_W,
            num_channels=int(base_config.num_channels),
        )

    mod_fmt = _modulation_format_for_model(model_key, spec)
    if spec["spectrum_type"] == SpectrumType.OSA_SAMPLED:
        band_limit_hz = 2.0 * float(base_config.B_s)  # 2 × symbol rate
        return build_wdm_grid(
            config=model_config,
            spectrum_type=spec["spectrum_type"],
            f_grid=f_grid,
            osa_csv_path=osa_csv_path,
            osa_rbw=_get_osa_rbw(osa_csv_path),
            osa_center_freq_hz=osa_center_freq_hz,
            osa_band_limit_hz=band_limit_hz,
            classical_channel_indices=CLASSICAL_INDICES,
            modulation_format=mod_fmt,
        )
    return build_wdm_grid(
        config=model_config,
        spectrum_type=spec["spectrum_type"],
        f_grid=f_grid,
        classical_channel_indices=CLASSICAL_INDICES,
        modulation_format=mod_fmt,
    )


def _display_channel_label(itn: int) -> str:
    """Convert ITU channel number to display label.

    Parameters
    ----------
    itn : int
        1-based ITU G.694.1 channel number (e.g. 39 = C39)
    """
    return f"C{itn}"


def adaptive_log_ticks(
    y_bot_log: float, y_top_log: float, max_ticks: int = 8
) -> dict:
    return dict(
        tickmode="auto",
        nticks=max_ticks,
        tickformat="1e",
        exponentformat="power",
    )


def adaptive_linear_ticks(
    y_bot: float, y_top: float, max_ticks: int = 8
) -> dict:
    return dict(
        tickmode="auto",
        nticks=max_ticks,
        tickformat=".2f",
    )


def _build_all_classical_grid(
    model_key: str,
    spec: dict,
    base_config: WDMConfig,
    f_grid: np.ndarray,
    osa_csv_path: Path,
    osa_center_freq_hz: float | None = None,
):
    from qkd_sim.physical.signal import SpectrumType

    model_config = WDMConfig(
        start_freq=base_config.start_freq,
        start_channel=base_config.start_channel,
        end_channel=base_config.end_channel,
        channel_spacing=base_config.channel_spacing,
        B_s=base_config.B_s,
        data_rate_bps=base_config.data_rate_bps,
        P0=_get_P0(),
        beta_rolloff=base_config.beta_rolloff if spec["beta_rolloff"] is None else spec["beta_rolloff"],
        ook_filter_order=base_config.ook_filter_order,
        quantum_channel_indices=base_config.quantum_channel_indices,
        channel_powers_W=base_config.channel_powers_W,
        num_channels=int(base_config.num_channels),
    )
    mod_fmt = _modulation_format_for_model(model_key, spec)
    # Do NOT pass classical_channel_indices — derive it as complement of quantum_channel_indices
    if spec["spectrum_type"] == SpectrumType.OSA_SAMPLED:
        band_limit_hz = 2.0 * float(base_config.B_s)  # 2 × symbol rate
        return build_wdm_grid(
            config=model_config,
            spectrum_type=spec["spectrum_type"],
            f_grid=f_grid,
            osa_csv_path=osa_csv_path,
            osa_rbw=_get_osa_rbw(osa_csv_path),
            osa_center_freq_hz=osa_center_freq_hz,
            osa_band_limit_hz=band_limit_hz,
            modulation_format=mod_fmt,
        )
    return build_wdm_grid(
        config=model_config,
        spectrum_type=spec["spectrum_type"],
        f_grid=f_grid,
        modulation_format=mod_fmt,
    )


def _make_fiber(fiber_params: dict, length_km: float):
    from qkd_sim.config.schema import FiberConfig
    from qkd_sim.physical.fiber import Fiber

    params = dict(fiber_params)
    params["L_km"] = float(length_km)
    return Fiber(FiberConfig(**params))


def _integrate_signal_per_channel(grid, f_grid: np.ndarray | None) -> np.ndarray:
    powers = np.zeros(len(grid.channels), dtype=np.float64)
    if f_grid is None or len(f_grid) < 2:
        for idx, ch in enumerate(grid.channels):
            if ch.channel_type == "classical":
                powers[idx] = float(ch.power)
        return powers

    df = float(np.mean(np.diff(f_grid)))
    from qkd_sim.physical.signal import SpectrumType
    for idx, ch in enumerate(grid.channels):
        if ch.channel_type == "classical":
            if ch.spectrum_type == SpectrumType.SINGLE_FREQ:
                # Delta approximation: sum(G*df) = P at nearest bin → G = P/df
                powers[idx] = float(ch.power)
            else:
                powers[idx] = float(np.sum(ch.get_psd(f_grid)) * df)
    return powers


def _integrate_noise_psd_per_channel(
    psd_fwd: np.ndarray,
    psd_bwd: np.ndarray,
    grid,
    f_grid: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:
    """积分连续 PSD 到各量子信道带宽。

    compute_forward_conti 返回完整 PSD (N_f,)，需积分到每个量子信道带宽
    得到各量子信道的积分噪声功率 (N_q,)。

    Parameters
    ----------
    psd_fwd, psd_bwd : ndarray (N_f,)
        前向/后向噪声 PSD
    grid : WDMGrid
    f_grid : ndarray (N_f,)

    Returns
    -------
    (N_q,) 前向/后向积分噪声功率
    """
    quantum_chs = grid.get_quantum_channels()
    n_q = len(quantum_chs)
    df = float(np.mean(np.diff(f_grid)))
    fwd_integrated = np.zeros(n_q, dtype=np.float64)
    bwd_integrated = np.zeros(n_q, dtype=np.float64)

    for i, ch in enumerate(quantum_chs):
        f_lo = ch.f_center - ch.B_s / 2.0
        f_hi = ch.f_center + ch.B_s / 2.0
        mask = (f_grid >= f_lo) & (f_grid < f_hi)
        fwd_integrated[i] = float(np.sum(psd_fwd[mask]) * df)
        bwd_integrated[i] = float(np.sum(psd_bwd[mask]) * df)

    return fwd_integrated, bwd_integrated


def _compute_noise_power_pair(
    noise_type: str,
    fiber,
    grid,
    continuous: bool,
    fwm_solver=None,
    sprs_solver=None,
    sprs_noise_bandwidth_hz: float | None = None,
) -> tuple[np.ndarray, np.ndarray]:
    """计算 FWM/SpRS 噪声积分功率 (N_q,)。

    连续模型：调用 compute_forward_conti/compute_backward_conti，返回 (N_q,) 积分功率
    离散模型：调用 compute_forward/compute_backward，返回 (N_q,) 积分功率

    Args:
        noise_type: "fwm" | "sprs" | "both"
        fiber: Fiber instance
        grid: WDMGrid
        continuous: 是否为连续模型
        fwm_solver: pre-constructed DiscreteFWMSolver (created once, reused across calls)
        sprs_solver: pre-constructed DiscreteSPRSSolver (created once, reused across calls)

    Returns
    -------
    (fwd, bwd), each shape (N_q,) — per-channel integrated noise power [W]
    """
    from qkd_sim.physical.noise import DiscreteFWMSolver, DiscreteSPRSSolver

    n_q = len(grid.get_quantum_channels())
    fwd = np.zeros(n_q, dtype=np.float64)
    bwd = np.zeros(n_q, dtype=np.float64)
    _f_grid = grid.f_grid

    def _call_solver(solver):
        if continuous:
            # compute_forward_conti 返回 (N_q,) 积分功率，不是 (N_f,)
            return (
                np.asarray(solver.compute_forward_conti(fiber, grid, _f_grid), dtype=np.float64),
                np.asarray(solver.compute_backward_conti(fiber, grid, _f_grid), dtype=np.float64),
            )
        # 离散模型：返回 (N_q,) 积分功率
        return (
            np.asarray(solver.compute_forward(fiber, grid), dtype=np.float64),
            np.asarray(solver.compute_backward(fiber, grid), dtype=np.float64),
        )

    if noise_type in ("fwm", "both"):
        solver = fwm_solver if fwm_solver is not None else DiscreteFWMSolver(active_threshold_db=ACTIVE_THRESHOLD_DB)
        f_i, b_i = _call_solver(solver)
        fwd += f_i
        bwd += b_i

    if noise_type in ("sprs", "both"):
        solver = sprs_solver if sprs_solver is not None else DiscreteSPRSSolver(
            noise_bandwidth_hz=sprs_noise_bandwidth_hz,
        )
        f_i, b_i = _call_solver(solver)
        fwd += f_i
        bwd += b_i

    return fwd, bwd


def _compute_noise_spectrum_pair(
    noise_type: str,
    fiber,
    grid,
    f_grid: np.ndarray,
    L_arr: np.ndarray | None = None,
    fwm_solver=None,
    sprs_solver=None,
    sprs_noise_bandwidth_hz: float | None = None,
) -> tuple[np.ndarray, np.ndarray]:
    """返回完整 PSD 数组，适用于连续模型。

    当 L_arr=None 时返回 (N_f,)，当 L_arr 给出时返回 (N_f, N_L) —
    每列对应一个光纤长度。

    fwm: compute_fwm_spectrum_conti
    sprs: compute_sprs_spectrum_conti
    both: 两者逐点相加

    fwm_solver / sprs_solver: pre-constructed solver instances (created once
    and reused across model_keys to preserve internal caches).

    Returns
    -------
    (fwd, bwd), each shape (N_f,) or (N_f, N_L)
    """
    from qkd_sim.physical.noise import DiscreteFWMSolver, DiscreteSPRSSolver

    n_f = len(f_grid)
    n_l = 1 if L_arr is None else int(L_arr.shape[0])
    fwd = np.zeros((n_f, n_l), dtype=np.float64)
    bwd = np.zeros((n_f, n_l), dtype=np.float64)

    if noise_type in ("fwm", "both"):
        solver = fwm_solver if fwm_solver is not None else DiscreteFWMSolver(active_threshold_db=ACTIVE_THRESHOLD_DB)
        fwm_fwd, fwm_bwd = solver.compute_fwm_spectrum_conti_pair(
            fiber, grid, f_grid, L_arr=L_arr
        )
        if L_arr is None:
            fwd[:, 0] += fwm_fwd
            bwd[:, 0] += fwm_bwd
        else:
            fwd += fwm_fwd
            bwd += fwm_bwd

    if noise_type in ("sprs", "both"):
        solver = sprs_solver if sprs_solver is not None else DiscreteSPRSSolver(
            noise_bandwidth_hz=sprs_noise_bandwidth_hz,
        )
        sprs_fwd, sprs_bwd = solver.compute_sprs_spectrum_conti(
            fiber, grid, f_grid, direction="both", L_arr=L_arr,
        )
        if L_arr is None:
            fwd[:, 0] += sprs_fwd
            bwd[:, 0] += sprs_bwd
        else:
            fwd += sprs_fwd
            bwd += sprs_bwd

    # Return (N_f,) for backward compatibility
    if n_l == 1:
        return np.asarray(fwd.ravel(), dtype=np.float64), np.asarray(bwd.ravel(), dtype=np.float64)
    return np.asarray(fwd, dtype=np.float64), np.asarray(bwd, dtype=np.float64)


def _integrate_psd_per_quantum_channel(
    psd: np.ndarray,
    f_grid: np.ndarray,
    q_center_freqs: np.ndarray,
    bandwidth: float,
    df: float,
) -> np.ndarray:
    """Integrate PSD over [f_q - B_q/2, f_q + B_q/2] for each quantum channel.

    Parameters
    ----------
    psd : ndarray, shape (N_f,) or (N_f, N_L)
        Noise PSD [W/Hz].
    f_grid : ndarray, shape (N_f,)
        Frequency grid [Hz].
    q_center_freqs : ndarray, shape (N_q,)
        Quantum channel center frequencies [Hz].
    bandwidth : float
        Integration bandwidth B_q [Hz].
    df : float
        Frequency grid spacing [Hz].

    Returns
    -------
    ndarray, shape (N_q,) or (N_q, N_L)
        Integrated noise power [W] per quantum channel.
    """
    n_q = len(q_center_freqs)
    scalar_input = psd.ndim == 1
    if scalar_input:
        psd = psd[:, np.newaxis]  # (N_f, 1)
    n_l = psd.shape[1]
    result = np.zeros((n_q, n_l), dtype=np.float64)
    half_bw = bandwidth / 2.0
    for qi, f_q in enumerate(q_center_freqs):
        in_band = (f_grid >= f_q - half_bw) & (f_grid <= f_q + half_bw)
        n_pts = np.sum(in_band)
        if n_pts == 0:
            continue
        if n_pts == 1:
            result[qi, :] = psd[in_band, :].ravel() * df
        else:
            result[qi, :] = np.trapezoid(psd[in_band, :], f_grid[in_band], axis=0)
    if scalar_input:
        return result.ravel()  # (N_q,)
    return result  # (N_q, N_L)


def precompute_by_length(
    noise_type: str,
    specs: dict,
    LENGTHS_KM: np.ndarray,
    base_config: WDMConfig,
    noise_f_grid: np.ndarray,
    osa_csv_path: Path,
    fiber_params: dict,
    osa_center_freq_hz: float | None = None,
) -> tuple[dict, list]:
    """Precompute noise vs fiber length for all quantum channels.

    Returns (ALL_BY_LEN, VALID_Q_INDICES)
    ALL_BY_LEN[q_idx][model_key] = {"fwd": np.array(N_L), "bwd": np.array(N_L)}
VALID_Q_INDICES: list of q_idx that have non-zero noise for at least one model
    """
    model_keys = get_noise_model_keys(noise_type)
    # Filter specs to only keys that are in model_keys (handles modulation format mismatch)
    specs = {k: specs[k] for k in model_keys if k in specs}
    quantum_indices = list(base_config.quantum_channel_indices)
    n_q = len(quantum_indices)
    n_l = len(LENGTHS_KM)
    n_ch = int(base_config.num_channels)
    power_label = _POWER_OVERRIDE_DBM if _POWER_OVERRIDE_DBM is not None else 0.0
    if _PROFILE_ENABLED:
        print(
            "[profile] precompute_by_channel "
            f"type={noise_type}, power={power_label:+.1f} dBm, "
            f"lengths={n_l}, quantum_channels={n_q}, freq_points={len(noise_f_grid)}, models={len(model_keys)}"
        )

    if noise_type == "with_signal":
        # Continuous PSD: noise (FWM+SpRS) + signal, integrate to power per length
        from qkd_sim.physical.signal import SpectrumType

        df = float(np.mean(np.diff(noise_f_grid)))
        all_by_len = {
            ch_idx: {mk: {
                "fwd": np.zeros(n_l), "bwd": np.zeros(n_l),
                "noise_only_fwd": np.zeros(n_l), "noise_only_bwd": np.zeros(n_l),
            } for mk in model_keys}
            for ch_idx in range(n_ch)
        }
        # Classical set: ITU G.694.1 channel numbers (e.g. {39, 40, 41} = C39/C40/C41)
        classical_set = set(CLASSICAL_INDICES)
        L_arr = np.array(LENGTHS_KM, dtype=np.float64) * 1e3  # km → m
        fiber_base = _make_fiber(fiber_params, LENGTHS_KM[0])
        model_config_noise = WDMConfig(
            start_freq=base_config.start_freq,
            start_channel=base_config.start_channel,
            end_channel=base_config.end_channel,
            channel_spacing=base_config.channel_spacing,
            B_s=base_config.B_s,
            B_q=base_config.B_q,
            data_rate_bps=base_config.data_rate_bps,
            P0=_get_P0(),
            beta_rolloff=WDM_PARAMS["beta_rolloff"],
            ook_filter_order=base_config.ook_filter_order,
            quantum_channel_indices=list(base_config.quantum_channel_indices),
            channel_powers_W=base_config.channel_powers_W,
            num_channels=int(base_config.num_channels),
        )
        grid_noise = build_wdm_grid(
            config=model_config_noise,
            spectrum_type=SpectrumType.RAISED_COSINE,
            f_grid=noise_f_grid,
            classical_channel_indices=CLASSICAL_INDICES,
            modulation_format="DP-16QAM",
        )
        # Vectorize over all lengths at once
        noise_fwd_psd_all, noise_bwd_psd_all = _compute_noise_spectrum_pair(
            "both", fiber_base, grid_noise, noise_f_grid, L_arr=L_arr,
            sprs_noise_bandwidth_hz=base_config.B_q,
        )  # shape (N_f, N_L) or (N_f,) when n_l==1
        # Guard: if n_l==1, result is 1D; expand to 2D for [:, li] indexing
        if noise_fwd_psd_all.ndim == 1:
            noise_fwd_psd_all = noise_fwd_psd_all[:, np.newaxis]
            noise_bwd_psd_all = noise_bwd_psd_all[:, np.newaxis]
        # Signal PSD: sum ONLY over classical channels (C39/C40/C41) — computed once, shared by all model_keys
        signal_psd = np.zeros(len(noise_f_grid), dtype=np.float64)
        for idx, ch in enumerate(grid_noise.channels):
            # idx is zero-based position; ITU channel number = idx+1
            if (idx + 1) in classical_set:
                signal_psd += ch.get_psd(noise_f_grid)
        # Quantum channel center frequencies and per-channel noise integration
        q_itn_list = list(base_config.quantum_channel_indices)
        q_center_freqs_len = np.array(
            [base_config.start_freq + (itn - base_config.start_channel) * base_config.channel_spacing
             for itn in q_itn_list],
            dtype=np.float64,
        )
        B_q_len = float(base_config.B_q)
        noise_only_fwd_by_q = _integrate_psd_per_quantum_channel(
            noise_fwd_psd_all, noise_f_grid, q_center_freqs_len, B_q_len, df,
        )  # (N_q, N_L)
        noise_only_bwd_by_q = _integrate_psd_per_quantum_channel(
            noise_bwd_psd_all, noise_f_grid, q_center_freqs_len, B_q_len, df,
        )  # (N_q, N_L)
        for model_key in model_keys:
            for li in range(n_l):
                noise_fwd_psd = noise_fwd_psd_all[:, li]
                noise_bwd_psd = noise_bwd_psd_all[:, li]
                total_fwd_power = float(np.sum(noise_fwd_psd + signal_psd) * df)
                total_bwd_power = float(np.sum(noise_bwd_psd + signal_psd) * df)
                for ch_idx in range(n_ch):
                    all_by_len[ch_idx][model_key]["fwd"][li] = total_fwd_power
                    all_by_len[ch_idx][model_key]["bwd"][li] = total_bwd_power
                # Per-channel noise: integrate PSD over B_q bandwidth per quantum channel
                for qi, itn in enumerate(q_itn_list):
                    ch_idx = int(itn) - 1  # ITU → 0-based
                    all_by_len[ch_idx][model_key]["noise_only_fwd"][li] = float(noise_only_fwd_by_q[qi, li])
                    all_by_len[ch_idx][model_key]["noise_only_bwd"][li] = float(noise_only_bwd_by_q[qi, li])

        valid_indices = [
            ch_idx
            for ch_idx in range(n_ch)
            if any(
                np.any(all_by_len[ch_idx][mk]["fwd"] > 0)
                or np.any(all_by_len[ch_idx][mk]["bwd"] > 0)
                for mk in model_keys
            )
        ]
        return all_by_len, valid_indices

    if noise_type == "only_signal":
        all_by_len = {
            ch_idx: {mk: {"fwd": np.zeros(n_l), "bwd": np.zeros(n_l)} for mk in model_keys}
            for ch_idx in range(n_ch)
        }
        for model_key in model_keys:
            spec = specs[model_key]
            grid_all = _build_all_classical_grid(model_key, spec, base_config, noise_f_grid, osa_csv_path, osa_center_freq_hz)
            signal = _integrate_signal_per_channel(grid_all, grid_all.f_grid)
            classical_mask = np.array([ch.channel_type == "classical" for ch in grid_all.channels], dtype=bool)

            # G_TX is independent of fiber length; compute once outside the length loop
            fwd = np.zeros(n_ch, dtype=np.float64)
            bwd = np.zeros(n_ch, dtype=np.float64)
            fwd[classical_mask] = signal[classical_mask]
            # bwd remains zeros (no backward-propagating signal at transmit side)

            for li in range(n_l):
                for ch_idx in range(n_ch):
                    all_by_len[ch_idx][model_key]["fwd"][li] = float(fwd[ch_idx])
                    all_by_len[ch_idx][model_key]["bwd"][li] = float(bwd[ch_idx])

        valid_indices = [
            ch_idx
            for ch_idx in range(n_ch)
            if any(
                np.any(all_by_len[ch_idx][mk]["fwd"] > 0)
                or np.any(all_by_len[ch_idx][mk]["bwd"] > 0)
                for mk in model_keys
            )
        ]
        return all_by_len, valid_indices

    # --- CSV caching for precomputed data ---
    _PRECOMPUTED_DIR = _PROJECT_ROOT / "data" / "precomputed"


    all_by_len = {
        q_local_idx: {mk: {"fwd": np.zeros(n_l), "bwd": np.zeros(n_l)} for mk in model_keys}
        for q_local_idx in range(n_q)
    }
    # Pre-build all grids per (q_local_idx, model_key) to avoid repeated _build_model_grid calls
    grid_cache: dict[tuple[int, str], object] = {}
    for q_local_idx, q_idx in enumerate(quantum_indices):
        single_q_config = _build_wdm_config([q_idx])
        for model_key in model_keys:
            spec = specs[model_key]
            cache_key = (q_local_idx, model_key)
            if cache_key not in grid_cache:
                grid_cache[cache_key] = _build_model_grid(
                    model_key,
                    spec,
                    single_q_config,
                    noise_f_grid,
                    osa_csv_path,
                    osa_center_freq_hz,
                )
            grid = grid_cache[cache_key]
            # Fiber varies with length, not with q_idx or model_key; create once per length
            for li, length_km in enumerate(LENGTHS_KM):
                fiber = _make_fiber(fiber_params, length_km)
                fwd, bwd = _compute_noise_power_pair(
                    noise_type,
                    fiber,
                    grid,
                    continuous=bool(spec["continuous"]),
                    sprs_noise_bandwidth_hz=base_config.B_q,
                )
                if len(fwd) > 0:
                    # single_q_config contains exactly one quantum channel, so the
                    # solver output is length-1 and should always be read at index 0.
                    all_by_len[q_local_idx][model_key]["fwd"][li] = float(fwd[0])
                    all_by_len[q_local_idx][model_key]["bwd"][li] = float(bwd[0])

    valid_q_indices = [
        q_local_idx
        for q_local_idx in range(n_q)
        if any(
            np.any(all_by_len[q_local_idx][mk]["fwd"] > 0)
            or np.any(all_by_len[q_local_idx][mk]["bwd"] > 0)
            for mk in model_keys
        )
    ]
    return all_by_len, valid_q_indices


def precompute_by_channel(
    noise_type: str,
    specs: dict,
    LENGTHS_KM: np.ndarray,
    base_config: WDMConfig,
    noise_f_grid: np.ndarray,
    osa_csv_path: Path,
    fiber_params: dict,
    osa_center_freq_hz: float | None = None,
) -> tuple[dict, list]:
    """Precompute noise vs quantum channel frequency for all lengths.

    Returns (ALL_BY_CH, VALID_L_INDICES)

    ALL_BY_CH[L_idx][model_key] = {
        "fwd": np.ndarray,       # (N_f,) for continuous PSD, (N_q,) for discrete channel power
        "bwd": np.ndarray,
        "x": np.ndarray,         # noise_f_grid for continuous, channel center freqs for discrete
        "x_kind": str,           # "frequency_grid" | "channel_center"
        "y_kind": str,           # "psd" | "channel_power"
    }
    VALID_L_INDICES: list of L indices with non-zero noise
    """
    model_keys = get_noise_model_keys(noise_type)
    # Filter specs to only keys that are in model_keys (handles modulation format mismatch)
    specs = {k: specs[k] for k in model_keys if k in specs}
    quantum_indices = list(base_config.quantum_channel_indices)
    n_q = len(quantum_indices)
    n_l = len(LENGTHS_KM)
    n_ch = int(base_config.num_channels)

    if noise_type == "with_signal":
        # Continuous PSD: noise (FWM+SpRS) + signal, output (N_f,) per length
        from qkd_sim.physical.signal import SpectrumType
        from qkd_sim.physical.noise import DiscreteFWMSolver, DiscreteSPRSSolver

        df = float(np.mean(np.diff(noise_f_grid)))
        # Quantum channel center frequencies for per-channel noise integration
        q_center_freqs = np.array(
            [base_config.start_freq + (itn - base_config.start_channel) * base_config.channel_spacing
             for itn in quantum_indices],
            dtype=np.float64,
        )
        B_q = float(base_config.B_q)
        all_by_ch = {
            li: {mk: {
                "fwd": np.zeros(len(noise_f_grid)),
                "bwd": np.zeros(len(noise_f_grid)),
                "noise_only_fwd": np.zeros(n_q),
                "noise_only_bwd": np.zeros(n_q),
                "noise_only_x": q_center_freqs,
                "x": np.asarray(noise_f_grid, dtype=np.float64),
                "x_kind": "frequency_grid",
                "y_kind": "power_per_bin",
            } for mk in model_keys}
            for li in range(n_l)
        }
        # Classical set: ITU G.694.1 channel numbers (e.g. {39, 40, 41})
        classical_set = set(CLASSICAL_INDICES)
        # Noise spectrum: use REAL quantum/classical split (only classical channels are pumps)
        # model_config_noise, grid_model, and signal_psd don't depend on fiber length —
        # build once per model_key to avoid redundant CSV loads and PSD integrations (T3)
        model_config_noise = WDMConfig(
            start_freq=base_config.start_freq,
            start_channel=base_config.start_channel,
            end_channel=base_config.end_channel,
            channel_spacing=base_config.channel_spacing,
            B_s=base_config.B_s,
            data_rate_bps=base_config.data_rate_bps,
            P0=_get_P0(),
            beta_rolloff=WDM_PARAMS["beta_rolloff"],
            ook_filter_order=base_config.ook_filter_order,
            quantum_channel_indices=list(base_config.quantum_channel_indices),
            channel_powers_W=base_config.channel_powers_W,
            num_channels=int(base_config.num_channels),
        )
        grid_cache: dict = {}
        signal_psd_cache: dict = {}
        for model_key in model_keys:
            spec = specs[model_key]
            if model_key == "discrete":
                stype = SpectrumType.SINGLE_FREQ
            elif model_key == "nrz_ook":
                stype = SpectrumType.NRZ_OOK
            elif model_key == "osa_ook":
                stype = SpectrumType.OSA_SAMPLED
            else:
                stype = SpectrumType.RAISED_COSINE
            beta_rolloff = (
                spec["beta_rolloff"]
                if spec.get("beta_rolloff") is not None
                else WDM_PARAMS["beta_rolloff"]
            )
            model_config = WDMConfig(
                start_freq=base_config.start_freq,
                start_channel=base_config.start_channel,
                end_channel=base_config.end_channel,
                channel_spacing=base_config.channel_spacing,
                B_s=base_config.B_s,
                data_rate_bps=base_config.data_rate_bps,
                P0=_get_P0(),
                beta_rolloff=beta_rolloff,
                ook_filter_order=base_config.ook_filter_order,
                quantum_channel_indices=list(base_config.quantum_channel_indices),
                channel_powers_W=base_config.channel_powers_W,
                num_channels=int(base_config.num_channels),
            )
            grid_kwargs = dict(
                config=model_config,
                spectrum_type=stype,
                f_grid=noise_f_grid,
                classical_channel_indices=CLASSICAL_INDICES,
                modulation_format=MODULATION_FORMAT,
            )
            if stype == SpectrumType.OSA_SAMPLED:
                grid_kwargs["osa_csv_path"] = osa_csv_path
                grid_kwargs["osa_rbw"] = _get_osa_rbw(osa_csv_path)
                grid_kwargs["osa_band_limit_hz"] = 2.0 * float(model_config.B_s)
            grid_model = build_wdm_grid(**grid_kwargs)
            grid_cache[model_key] = grid_model
            # Signal PSD: sum ONLY over classical channels (C39/C40/C41)
            signal_psd = np.zeros_like(noise_f_grid, dtype=np.float64)
            for idx, ch in enumerate(grid_model.channels):
                if (idx + 1) in classical_set:
                    signal_psd += ch.get_psd(noise_f_grid)
            signal_psd_cache[model_key] = signal_psd

        _ws_fwm = DiscreteFWMSolver(active_threshold_db=ACTIVE_THRESHOLD_DB)
        _ws_sprs = DiscreteSPRSSolver(noise_bandwidth_hz=B_q)

        for li, length_km in enumerate(LENGTHS_KM):
            fiber = _make_fiber(fiber_params, length_km)
            for model_key in model_keys:
                grid_model = grid_cache[model_key]
                noise_fwd_psd, noise_bwd_psd = _compute_noise_spectrum_pair(
                    "both", fiber, grid_model, noise_f_grid,
                    fwm_solver=_ws_fwm, sprs_solver=_ws_sprs,
                )
                signal_psd = signal_psd_cache[model_key]
                total_fwd = (noise_fwd_psd + signal_psd) * df
                total_bwd = (noise_bwd_psd + signal_psd) * df
                all_by_ch[li][model_key]["fwd"] = np.asarray(total_fwd, dtype=np.float64)
                all_by_ch[li][model_key]["bwd"] = np.asarray(total_bwd, dtype=np.float64)
                # T1: integrate pure noise PSD over B_q bandwidth per quantum channel
                all_by_ch[li][model_key]["noise_only_fwd"] = _integrate_psd_per_quantum_channel(
                    noise_fwd_psd, noise_f_grid, q_center_freqs, B_q, df,
                )
                all_by_ch[li][model_key]["noise_only_bwd"] = _integrate_psd_per_quantum_channel(
                    noise_bwd_psd, noise_f_grid, q_center_freqs, B_q, df,
                )

        valid_l_indices = [
            li
            for li in range(n_l)
            if any(
                np.any(all_by_ch[li][mk]["fwd"] > 0)
                or np.any(all_by_ch[li][mk]["bwd"] > 0)
                for mk in model_keys
            )
        ]
        return all_by_ch, valid_l_indices

    if noise_type == "only_signal":
        # Continuous models: use full frequency grid + PSD for waveform display
        # Discrete model: use channel power at classical channel centers
        all_by_ch: dict = {li: {} for li in range(n_l)}
        df = float(np.mean(np.diff(noise_f_grid)))
        for model_key in model_keys:
            spec = specs[model_key]
            grid_all = _build_all_classical_grid(model_key, spec, base_config, noise_f_grid, osa_csv_path, osa_center_freq_hz)
            classical_mask = np.array(
                [ch.channel_type == "classical" for ch in grid_all.channels], dtype=bool
            )

            if spec["continuous"]:
                # Continuous model: full f_grid + PSD → waveform line
                psd_total = grid_all.get_total_psd()  # shape (N_f,)
                fwd_arr = np.asarray(psd_total * df, dtype=np.float64)
                bwd_arr = np.zeros(len(psd_total), dtype=np.float64)
                x_arr = np.asarray(noise_f_grid, dtype=np.float64)
                x_kind = "frequency_grid"
                y_kind = "power_per_bin"
            else:
                # Discrete model: channel power at classical channel centers
                signal = _integrate_signal_per_channel(grid_all, noise_f_grid)
                fwd = np.zeros(n_ch, dtype=np.float64)
                fwd[classical_mask] = signal[classical_mask]
                fwd_arr = np.asarray(fwd, dtype=np.float64)
                bwd_arr = np.zeros(n_ch, dtype=np.float64)
                x_arr = np.array(
                    [WDM_PARAMS["start_freq"] + idx * WDM_PARAMS["channel_spacing"]
                     for idx in range(n_ch)],
                    dtype=np.float64,
                )
                x_kind = "channel_center"
                y_kind = "channel_power"

            for li in range(n_l):
                all_by_ch[li][model_key] = {
                    "fwd": fwd_arr,
                    "bwd": bwd_arr,
                    "x": x_arr,
                    "x_kind": x_kind,
                    "y_kind": y_kind,
                }

        valid_l_indices = [
            li
            for li in range(n_l)
            if any(
                np.any(all_by_ch[li][mk]["fwd"] > 0)
                or np.any(all_by_ch[li][mk]["bwd"] > 0)
                for mk in model_keys
            )
        ]
        return all_by_ch, valid_l_indices

    # Precompute quantum channel center frequencies for discrete-model x-axis
    q_center_freqs = np.array(
        [
            WDM_PARAMS["start_freq"] + (idx - WDM_PARAMS["start_channel"]) * WDM_PARAMS["channel_spacing"]
            for idx in quantum_indices
        ],
        dtype=np.float64,
    )

    # Initialize all_by_ch with metadata structure
    all_by_ch: dict = {}
    for li in range(n_l):
        all_by_ch[li] = {}
        for mk in model_keys:
            all_by_ch[li][mk] = {
                "fwd": np.array([]),
                "bwd": np.array([]),
                "x": np.array([]),
                "x_kind": "",
                "y_kind": "",
            }

    # --- Multiprocessing dispatch for default noise types ---
    if _MP_ENABLED and n_l > 1 and noise_type not in ("with_signal", "only_signal"):
        # Import worker from separate module (avoids Windows spawn path issues)
        from scripts.dash_workers import mp_worker_single_length

        # Serialize data for subprocess
        wdm_cfg_dict = dict(
            start_freq=base_config.start_freq,
            start_channel=base_config.start_channel,
            end_channel=base_config.end_channel,
            channel_spacing=base_config.channel_spacing,
            B_s=base_config.B_s,
            data_rate_bps=base_config.data_rate_bps,
            P0=_get_P0(),
            beta_rolloff=WDM_PARAMS["beta_rolloff"],
            ook_filter_order=base_config.ook_filter_order,
            quantum_channel_indices=list(base_config.quantum_channel_indices),
            num_channels=int(base_config.num_channels),
        )
        noise_f_list = noise_f_grid.tolist()
        power_dbm = _POWER_OVERRIDE_DBM

        futures = {}
        with ProcessPoolExecutor(max_workers=_get_mp_workers()) as executor:
            for li, length_km in enumerate(LENGTHS_KM):
                fut = executor.submit(
                    mp_worker_single_length,
                    li, length_km, noise_type,
                    model_keys, specs,
                    wdm_cfg_dict, noise_f_list,
                    fiber_params,
                    CLASSICAL_INDICES,
                    len(noise_f_list),
                    power_dbm,
                )
                futures[fut] = li

            for fut in as_completed(futures):
                li, results = fut.result()
                for mk, res in zip(model_keys, results):
                    all_by_ch[li][mk] = res

        valid_l_indices = [
            li for li in range(n_l)
            if any(
                np.any(all_by_ch[li][mk]["fwd"] > 0)
                or np.any(all_by_ch[li][mk]["bwd"] > 0)
                for mk in model_keys
            )
        ]
        return all_by_ch, valid_l_indices
    # --- End multiprocessing ---

    # Continuous model: iterate per length (original approach, avoids L_arr issues with sprs)
    # Cache: grid depends only on model_key (not on q_idx or li)
    grid_cache: dict[str, object] = {}

    # Create solver instances once for continuous models (shared across model_keys
    # to preserve SpRS sigma cache and FWM topology cache).
    from qkd_sim.physical.noise import DiscreteFWMSolver, DiscreteSPRSSolver
    _cont_fwm = DiscreteFWMSolver(active_threshold_db=ACTIVE_THRESHOLD_DB) if noise_type in ("fwm", "both") else None
    _cont_sprs = DiscreteSPRSSolver() if noise_type in ("sprs", "both") else None

    for model_key in model_keys:
        spec = specs[model_key]
        if not spec["continuous"]:
            continue
        t_grid = 0.0
        t_fiber = 0.0
        t_solver = 0.0
        t_store = 0.0
        if model_key not in grid_cache:
            _t = time.perf_counter()
            grid_cache[model_key] = _build_model_grid(
                model_key,
                spec,
                base_config,
                noise_f_grid,
                osa_csv_path,
                osa_center_freq_hz,
            )
            t_grid += time.perf_counter() - _t
        grid = grid_cache[model_key]
        df = float(np.mean(np.diff(noise_f_grid)))

        if noise_type in ("fwm", "sprs"):
            _t = time.perf_counter()
            fiber = _make_fiber(fiber_params, float(LENGTHS_KM[0]))
            L_arr = np.asarray(LENGTHS_KM, dtype=np.float64) * 1e3
            t_fiber += time.perf_counter() - _t
            _t = time.perf_counter()
            fwd_psd_all, bwd_psd_all = _compute_noise_spectrum_pair(
                noise_type, fiber, grid, noise_f_grid, L_arr=L_arr,
                fwm_solver=_cont_fwm, sprs_solver=_cont_sprs,
            )  # shape (N_f, N_L)
            t_solver += time.perf_counter() - _t
            if fwd_psd_all.ndim == 1:
                fwd_psd_all = fwd_psd_all[:, np.newaxis]
                bwd_psd_all = bwd_psd_all[:, np.newaxis]
            for li in range(n_l):
                _t = time.perf_counter()
                all_by_ch[li][model_key] = {
                    "fwd": np.asarray(fwd_psd_all[:, li] * df, dtype=np.float64),
                    "bwd": np.asarray(bwd_psd_all[:, li] * df, dtype=np.float64),
                    "x": np.asarray(noise_f_grid, dtype=np.float64),
                    "x_kind": "frequency_grid",
                    "y_kind": "power_per_bin",
                }
                t_store += time.perf_counter() - _t
        else:
            for li, length_km in enumerate(LENGTHS_KM):
                _t = time.perf_counter()
                fiber = _make_fiber(fiber_params, length_km)
                t_fiber += time.perf_counter() - _t
                _t = time.perf_counter()
                fwd_psd, bwd_psd = _compute_noise_spectrum_pair(
                    noise_type, fiber, grid, noise_f_grid,
                    fwm_solver=_cont_fwm, sprs_solver=_cont_sprs,
                )  # shape (N_f,)
                t_solver += time.perf_counter() - _t
                _t = time.perf_counter()
                all_by_ch[li][model_key] = {
                    "fwd": np.asarray(fwd_psd * df, dtype=np.float64),
                    "bwd": np.asarray(bwd_psd * df, dtype=np.float64),
                    "x": np.asarray(noise_f_grid, dtype=np.float64),
                    "x_kind": "frequency_grid",
                    "y_kind": "power_per_bin",
                }
                t_store += time.perf_counter() - _t
        if _PROFILE_ENABLED:
            print(
                "[profile]   continuous "
                f"model={model_key}: grid={t_grid:.3f}s, fiber={t_fiber:.3f}s, "
                f"solver={t_solver:.3f}s, store={t_store:.3f}s"
            )

    # Discrete model: build ONE grid with ALL quantum channels per model_key.
    # Both DiscreteFWMSolver and DiscreteSPRSSolver already support multi-
    # quantum-channel computation natively — calling once with N_q channels
    # produces the same result as N_q individual calls with 1 channel each.
    if any(not spec["continuous"] for spec in specs.values()):
        from qkd_sim.physical.noise import DiscreteFWMSolver, DiscreteSPRSSolver
        _disc_fwm = DiscreteFWMSolver(active_threshold_db=ACTIVE_THRESHOLD_DB) if noise_type in ("fwm", "both") else None
        _disc_sprs = DiscreteSPRSSolver() if noise_type in ("sprs", "both") else None
        # Build one multi-quantum grid with ALL quantum channels
        _multi_q_config = _build_wdm_config(quantum_indices)
    else:
        _disc_fwm = None
        _disc_sprs = None
        _multi_q_config = None

    for model_key in model_keys:
        spec = specs[model_key]
        if spec["continuous"]:
            continue  # Already handled above
        t_grid = 0.0
        t_solver = 0.0
        t_store = 0.0

        _t = time.perf_counter()
        grid_multi = _build_model_grid(
            model_key, spec, _multi_q_config,
            noise_f_grid, osa_csv_path, osa_center_freq_hz,
        )
        t_grid = time.perf_counter() - _t

        # Precompute SpRS for all lengths via L_arr (sigma computed once)
        _sprs_fwd_all = None
        _sprs_bwd_all = None
        if _disc_sprs is not None:
            _L_arr_m = np.asarray(LENGTHS_KM, dtype=np.float64) * 1e3
            _fiber_base = _make_fiber(fiber_params, float(LENGTHS_KM[0]))
            _sprs_fwd_all = _disc_sprs.compute_forward_l_array(_fiber_base, grid_multi, _L_arr_m)
            _sprs_bwd_all = _disc_sprs.compute_backward_l_array(_fiber_base, grid_multi, _L_arr_m)

        for li, length_km in enumerate(LENGTHS_KM):
            fiber = _make_fiber(fiber_params, length_km)
            _t = time.perf_counter()
            if _disc_sprs is not None and _disc_fwm is not None:
                # noise_type == "both": FWM per-length + SpRS from L_arr
                fwd_fwm, bwd_fwm = _compute_noise_power_pair(
                    "fwm", fiber, grid_multi, continuous=False,
                    fwm_solver=_disc_fwm,
                )
                fwd = fwd_fwm + _sprs_fwd_all[:, li]
                bwd = bwd_fwm + _sprs_bwd_all[:, li]
            elif _disc_sprs is not None:
                # noise_type == "sprs": from L_arr
                fwd = _sprs_fwd_all[:, li]
                bwd = _sprs_bwd_all[:, li]
            else:
                # noise_type == "fwm": per-length only
                fwd, bwd = _compute_noise_power_pair(
                    noise_type, fiber, grid_multi, continuous=False,
                    fwm_solver=_disc_fwm,
                )
            t_solver += time.perf_counter() - _t
            _t = time.perf_counter()
            all_by_ch[li][model_key] = {
                "fwd": np.asarray(fwd, dtype=np.float64),
                "bwd": np.asarray(bwd, dtype=np.float64),
                "x": np.asarray(q_center_freqs, dtype=np.float64),
                "x_kind": "channel_center",
                "y_kind": "channel_power",
            }
            t_store += time.perf_counter() - _t
        if _PROFILE_ENABLED:
            print(
                "[profile]   discrete "
                f"model={model_key}: grid={t_grid:.3f}s, "
                f"solver={t_solver:.3f}s, store={t_store:.3f}s, solver_calls={n_l}"
            )

    valid_l_indices = [
        li
        for li in range(n_l)
        if any(
            np.any(all_by_ch[li][mk]["fwd"] > 0)
            or np.any(all_by_ch[li][mk]["bwd"] > 0)
            for mk in model_keys
        )
    ]
    return all_by_ch, valid_l_indices


def get_noise_model_keys(noise_type: str) -> list[str]:
    """Return model keys for the given noise_type, respecting MODULATION_FORMAT."""
    from qkd_sim.config.plot_config import load_model_specs

    # FWM and SpRS share the same model group (same spectral models, different solvers)
    _mod_key = MODULATION_FORMAT.lower()
    if noise_type in ("fwm", "sprs"):
        group = f"fwm_noise_{_mod_key}"
        return _apply_model_key_filter(list(load_model_specs(group).keys()))

    # only_signal / with_signal / both: same models as fwm (signal + noise combined differently)
    # Return the fwm model keys (same spectral models, different computation path)
    group = f"fwm_noise_{_mod_key}"
    return _apply_model_key_filter(list(load_model_specs(group).keys()))


# --- Startup precomputation of ALL power levels (step=5 dBm, 7 values) ---
def precompute_by_channel_all_powers(
    noise_type: str,
    specs: dict,
    LENGTHS_KM: np.ndarray,
    base_config: WDMConfig,
    noise_f_grid: np.ndarray,
    osa_csv_path: Path,
    fiber_params: dict,
    osa_center_freq_hz: float | None = None,
) -> tuple[dict, list]:
    """Precompute ALL power levels at startup, save to CSV and memory cache.

    Clears stale CSV cache first, then computes for each power level in
    PRECOMPUTE_POWER_LEVELS in parallel (if total work justifies spawn overhead),
    stores in both CSV (for Origin) and _POWER_CACHE (for instant slider response).

    Returns (all_by_ch at 0 dBm, valid_l_indices).
    """
    import multiprocessing

    _print_gpu_status_once()

    with profile_scope("precompute_by_channel_all_powers: clear stale CSV cache"):
        _clear_precomputed_csv_files("ch")
        _POWER_CACHE.clear()

    if noise_type in ("sprs", "fwm"):
        print(
            "[profile] precompute_by_channel_all_powers: "
            f"using 0 dBm base result plus {'linear' if noise_type == 'sprs' else 'cubic'} power scaling"
        )
        set_power_override(0.0)
        with profile_scope("power +0 dBm: base precompute_by_channel"):
            base_all_by_idx, base_valid_l = precompute_by_channel(
                noise_type, specs, LENGTHS_KM, base_config,
                noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
            )
        results: dict[float, tuple[dict, list]] = {}
        for p in PRECOMPUTE_POWER_LEVELS:
            scale = _power_scaling_factor(noise_type, float(p))
            assert scale is not None
            with profile_scope(f"power {p:+.0f} dBm: scale from 0 dBm"):
                all_by_idx = (
                    base_all_by_idx
                    if float(p) == 0.0
                    else _scale_precomputed_result(base_all_by_idx, scale)
                )
            with profile_scope(f"power {p:+.0f} dBm: save CSV cache"):
                _cache_precomputed_result(all_by_idx, p, noise_type, list(specs.keys()), index_prefix="ch")
            _POWER_CACHE[("ch", float(p))] = all_by_idx
            results[float(p)] = (all_by_idx, base_valid_l)

        set_power_override(0.0)
        return results[0.0]

    if noise_type == "both":
        print(
            "[profile] precompute_by_channel_all_powers: "
            "using separate 0 dBm FWM/SpRS bases plus cubic/linear power scaling"
        )
        set_power_override(0.0)
        with profile_scope("power +0 dBm: base FWM precompute_by_channel"):
            fwm_base_by_idx, fwm_valid_l = precompute_by_channel(
                "fwm", specs, LENGTHS_KM, base_config,
                noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
            )
        with profile_scope("power +0 dBm: base SpRS precompute_by_channel"):
            sprs_base_by_idx, sprs_valid_l = precompute_by_channel(
                "sprs", specs, LENGTHS_KM, base_config,
                noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
            )
        base_valid_l = sorted(set(fwm_valid_l) & set(sprs_valid_l))
        results: dict[float, tuple[dict, list]] = {}
        for p in PRECOMPUTE_POWER_LEVELS:
            fwm_scale = _power_scaling_factor("fwm", float(p))
            sprs_scale = _power_scaling_factor("sprs", float(p))
            assert fwm_scale is not None and sprs_scale is not None
            with profile_scope(f"power {p:+.0f} dBm: combine FWM/SpRS bases"):
                all_by_idx = _combine_precomputed_results(
                    fwm_base_by_idx, sprs_base_by_idx, fwm_scale, sprs_scale
                )
            with profile_scope(f"power {p:+.0f} dBm: save CSV cache"):
                _cache_precomputed_result(all_by_idx, p, noise_type, list(specs.keys()), index_prefix="ch")
            _POWER_CACHE[("ch", float(p))] = all_by_idx
            results[float(p)] = (all_by_idx, base_valid_l)

        set_power_override(0.0)
        return results[0.0]

    if noise_type == "with_signal":
        print(
            "[profile] precompute_by_channel_all_powers: "
            "with_signal: FWM(P^3) + SpRS(P^1) + signal(P^1) scaling from 0 dBm base"
        )
        from qkd_sim.physical.signal import SpectrumType
        from qkd_sim.physical.noise import DiscreteFWMSolver, DiscreteSPRSSolver
        set_power_override(0.0)
        classical_set = set(CLASSICAL_INDICES)
        df = float(np.mean(np.diff(noise_f_grid)))
        model_config_noise = WDMConfig(
            start_freq=base_config.start_freq,
            start_channel=base_config.start_channel,
            end_channel=base_config.end_channel,
            channel_spacing=base_config.channel_spacing,
            B_s=base_config.B_s,
            data_rate_bps=base_config.data_rate_bps,
            P0=_get_P0(),
            beta_rolloff=WDM_PARAMS["beta_rolloff"],
            ook_filter_order=base_config.ook_filter_order,
            quantum_channel_indices=list(base_config.quantum_channel_indices),
            channel_powers_W=base_config.channel_powers_W,
            num_channels=int(base_config.num_channels),
        )
        model_keys_ws = get_noise_model_keys("with_signal")
        specs_ws = {k: specs[k] for k in model_keys_ws if k in specs}
        n_l_ws = len(LENGTHS_KM)
        cache_payload = _with_signal_cache_payload(
            noise_type,
            model_keys_ws,
            LENGTHS_KM,
            base_config,
            noise_f_grid,
            osa_csv_path,
            fiber_params,
            osa_center_freq_hz,
        )
        cached = _load_persistent_cache("ch_with_signal", cache_payload)
        if cached is not None:
            _POWER_CACHE.update(cached["power_cache"])
            set_power_override(0.0)
            return cached["result_all"], cached["valid_l"]

        # Build WDM grid and signal PSD once per model_key (independent of length and power)
        grid_cache_t4: dict = {}
        signal_base_cache: dict = {}  # model_key → signal_psd [W/Hz] at 0 dBm
        for mk in model_keys_ws:
            if mk not in specs_ws:
                continue
            spec_t4 = specs_ws[mk]
            if mk == "discrete":
                stype_t4 = SpectrumType.SINGLE_FREQ
            elif mk == "nrz_ook":
                stype_t4 = SpectrumType.NRZ_OOK
            elif mk in ("osa", "osa_ook"):
                stype_t4 = SpectrumType.OSA_SAMPLED
            else:
                stype_t4 = SpectrumType.RAISED_COSINE
            beta_rolloff_t4 = (
                spec_t4["beta_rolloff"]
                if spec_t4.get("beta_rolloff") is not None
                else WDM_PARAMS["beta_rolloff"]
            )
            model_config_t4 = WDMConfig(
                start_freq=base_config.start_freq,
                start_channel=base_config.start_channel,
                end_channel=base_config.end_channel,
                channel_spacing=base_config.channel_spacing,
                B_s=base_config.B_s,
                data_rate_bps=base_config.data_rate_bps,
                P0=_get_P0(),
                beta_rolloff=beta_rolloff_t4,
                ook_filter_order=base_config.ook_filter_order,
                quantum_channel_indices=list(base_config.quantum_channel_indices),
                channel_powers_W=base_config.channel_powers_W,
                num_channels=int(base_config.num_channels),
            )
            gkw_t4 = dict(
                config=model_config_t4,
                spectrum_type=stype_t4,
                f_grid=noise_f_grid,
                classical_channel_indices=CLASSICAL_INDICES,
                modulation_format=MODULATION_FORMAT,
            )
            if stype_t4 == SpectrumType.OSA_SAMPLED:
                gkw_t4["osa_csv_path"] = osa_csv_path
                gkw_t4["osa_rbw"] = _get_osa_rbw(osa_csv_path)
                gkw_t4["osa_center_freq_hz"] = osa_center_freq_hz
                gkw_t4["osa_band_limit_hz"] = 2.0 * float(model_config_t4.B_s)
            grid_t4 = build_wdm_grid(**gkw_t4)
            grid_cache_t4[mk] = grid_t4
            sig_base = np.zeros_like(noise_f_grid, dtype=np.float64)
            for idx, ch in enumerate(grid_t4.channels):
                if (idx + 1) in classical_set:
                    sig_base += ch.get_psd(noise_f_grid)
            signal_base_cache[mk] = sig_base  # [W/Hz] at 0 dBm

        # Compute FWM-only and SpRS-only PSDs at 0 dBm for each (li, model_key)
        with profile_scope("power +0 dBm: base FWM+SpRS precompute (with_signal)"):
            fwm_base_ws: dict = {
                li: {mk: {"fwd": np.zeros(len(noise_f_grid)), "bwd": np.zeros(len(noise_f_grid))} for mk in model_keys_ws}
                for li in range(n_l_ws)
            }
            sprs_base_ws: dict = {
                li: {mk: {"fwd": np.zeros(len(noise_f_grid)), "bwd": np.zeros(len(noise_f_grid))} for mk in model_keys_ws}
                for li in range(n_l_ws)
            }
            x_base_ws: dict = {mk: np.asarray(noise_f_grid, dtype=np.float64) for mk in model_keys_ws}
            L_arr_ws = np.asarray(LENGTHS_KM, dtype=np.float64) * 1e3  # (N_L,)
            fiber_t4 = _make_fiber(fiber_params, float(LENGTHS_KM[0]))  # L only for metadata; actual lengths via L_arr
            B_q_ws = float(base_config.B_q)
            _ws2_fwm = DiscreteFWMSolver(active_threshold_db=ACTIVE_THRESHOLD_DB)
            _ws2_sprs = DiscreteSPRSSolver(noise_bandwidth_hz=B_q_ws)

            for mk in model_keys_ws:
                if mk not in grid_cache_t4:
                    continue
                grid_t4 = grid_cache_t4[mk]

                with profile_scope(f"FWM L_arr batch [{mk}]"):
                    fwm_fwd_all, fwm_bwd_all = _compute_noise_spectrum_pair(
                        "fwm", fiber_t4, grid_t4, noise_f_grid, L_arr=L_arr_ws,
                        fwm_solver=_ws2_fwm, sprs_solver=_ws2_sprs,
                    )  # (N_f, N_L)
                with profile_scope(f"SpRS L_arr batch [{mk}]"):
                    sprs_fwd_all, sprs_bwd_all = _compute_noise_spectrum_pair(
                        "sprs", fiber_t4, grid_t4, noise_f_grid, L_arr=L_arr_ws,
                        fwm_solver=_ws2_fwm, sprs_solver=_ws2_sprs,
                    )  # (N_f, N_L)

                for li in range(n_l_ws):
                    fwm_base_ws[li][mk]["fwd"] = np.asarray(fwm_fwd_all[:, li] * df, dtype=np.float64)
                    fwm_base_ws[li][mk]["bwd"] = np.asarray(fwm_bwd_all[:, li] * df, dtype=np.float64)
                    sprs_base_ws[li][mk]["fwd"] = np.asarray(sprs_fwd_all[:, li] * df, dtype=np.float64)
                    sprs_base_ws[li][mk]["bwd"] = np.asarray(sprs_bwd_all[:, li] * df, dtype=np.float64)
        base_valid_l_ws = [
            li for li in range(n_l_ws)
            if any(
                np.any(fwm_base_ws[li][mk]["fwd"] > 0) or np.any(sprs_base_ws[li][mk]["fwd"] > 0)
                for mk in model_keys_ws if mk in fwm_base_ws[li]
            )
        ]

        # Quantum channel center frequencies and bandwidth for per-channel noise integration
        quantum_indices_ws = list(base_config.quantum_channel_indices)
        q_center_freqs_ws = np.array(
            [base_config.start_freq + (itn - base_config.start_channel) * base_config.channel_spacing
             for itn in quantum_indices_ws],
            dtype=np.float64,
        )
        B_q_ws = float(base_config.B_q)

        results: dict[float, tuple[dict, list]] = {}
        for p in PRECOMPUTE_POWER_LEVELS:
            fwm_scale = _power_scaling_factor("fwm", float(p))   # P³
            sprs_scale = _power_scaling_factor("sprs", float(p))  # P¹
            sig_scale = float(10.0 ** (float(p) / 10.0))          # P¹

            all_by_idx: dict = {}
            for li in range(n_l_ws):
                all_by_idx[li] = {}
                for mk in model_keys_ws:
                    if mk not in fwm_base_ws[li]:
                        continue
                    fwm_fwd = fwm_base_ws[li][mk]["fwd"]
                    fwm_bwd = fwm_base_ws[li][mk]["bwd"]
                    sprs_fwd = sprs_base_ws[li][mk]["fwd"]
                    sprs_bwd = sprs_base_ws[li][mk]["bwd"]
                    sig_psd = signal_base_cache.get(mk, np.zeros(len(noise_f_grid)))
                    sig_scaled = sig_scale * sig_psd * df  # [W/bin]

                    noise_only_fwd_bin = fwm_scale * fwm_fwd + sprs_scale * sprs_fwd
                    noise_only_bwd_bin = fwm_scale * fwm_bwd + sprs_scale * sprs_bwd
                    # Integrate per-bin noise over B_q bandwidth per quantum channel
                    noise_only_fwd = _integrate_psd_per_quantum_channel(
                        noise_only_fwd_bin / df, noise_f_grid, q_center_freqs_ws, B_q_ws, df,
                    )
                    noise_only_bwd = _integrate_psd_per_quantum_channel(
                        noise_only_bwd_bin / df, noise_f_grid, q_center_freqs_ws, B_q_ws, df,
                    )
                    all_by_idx[li][mk] = {
                        "fwd": noise_only_fwd_bin + sig_scaled,
                        "bwd": noise_only_bwd_bin + sig_scaled,
                        "noise_only_fwd": noise_only_fwd,
                        "noise_only_bwd": noise_only_bwd,
                        "noise_only_x": q_center_freqs_ws,
                        "x": x_base_ws[mk],
                        "x_kind": "frequency_grid",
                        "y_kind": "power_per_bin",
                    }

            with profile_scope(f"power {p:+.0f} dBm: save CSV cache"):
                _cache_precomputed_result(all_by_idx, p, noise_type, list(specs_ws.keys()), index_prefix="ch")
            _POWER_CACHE[("ch", float(p))] = all_by_idx
            results[float(p)] = (all_by_idx, base_valid_l_ws)

        set_power_override(0.0)
        _save_persistent_cache(
            "ch_with_signal",
            cache_payload,
            {
                "power_cache": {("ch", float(p)): results[float(p)][0] for p in PRECOMPUTE_POWER_LEVELS},
                "result_all": results[0.0][0],
                "valid_l": results[0.0][1],
            },
        )
        return results[0.0]

    # Sequential fast path: avoids spawn overhead for fast noise types (e.g. sprs ~1s total)
    if not _MP_ENABLED or len(PRECOMPUTE_POWER_LEVELS) <= 2:
        results: dict[float, tuple[dict, list]] = {}
        for p in PRECOMPUTE_POWER_LEVELS:
            set_power_override(float(p))
            with profile_scope(f"power {p:+.0f} dBm: precompute_by_channel"):
                all_by_idx, valid_l = precompute_by_channel(
                    noise_type, specs, LENGTHS_KM, base_config,
                    noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
                )
            with profile_scope(f"power {p:+.0f} dBm: save CSV cache"):
                _cache_precomputed_result(all_by_idx, p, noise_type, list(specs.keys()), index_prefix="ch")
            _POWER_CACHE[("ch", p)] = all_by_idx
            results[p] = (all_by_idx, valid_l)
    else:
        # Parallel: all power levels are independent
        n_workers = min(len(PRECOMPUTE_POWER_LEVELS), _get_mp_workers())
        ctx = multiprocessing.get_context()
        with profile_scope(f"all powers: parallel precompute ({n_workers} workers)"):
            with ProcessPoolExecutor(max_workers=n_workers, mp_context=ctx) as executor:
                futures = {
                    executor.submit(
                        _precompute_channel_worker,
                        p, noise_type, specs, LENGTHS_KM, base_config,
                        noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
                    ): p
                    for p in PRECOMPUTE_POWER_LEVELS
                }
                results = {}
                for fut in futures:
                    p, all_by_idx, valid_l = fut.result()
                    # Cache is already written inside the worker
                    _POWER_CACHE[("ch", p)] = all_by_idx
                    results[p] = (all_by_idx, valid_l)

    # Extract 0 dBm result
    result_all, result_valid = results[0.0]

    # Restore 0 dBm in memory cache and power override
    set_power_override(0.0)
    _POWER_CACHE[("ch", 0.0)] = result_all
    return result_all, result_valid


def precompute_by_length_all_powers(
    noise_type: str,
    specs: dict,
    LENGTHS_KM: np.ndarray,
    base_config: WDMConfig,
    noise_f_grid: np.ndarray,
    osa_csv_path: Path,
    fiber_params: dict,
    osa_center_freq_hz: float | None = None,
) -> tuple[dict, list]:
    """Precompute ALL power levels at startup, save to CSV and memory cache.

    Clears stale CSV cache first, then computes for each power level in
    PRECOMPUTE_POWER_LEVELS in parallel (if total work justifies spawn overhead),
    stores in both CSV (for Origin) and _POWER_CACHE (for instant slider response).

    Returns (all_by_len at 0 dBm, valid_ch_indices).
    """
    import multiprocessing

    _print_gpu_status_once()

    _clear_precomputed_csv_files("len")
    _POWER_CACHE.clear()

    if noise_type == "both":
        print(
            "[profile] precompute_by_length_all_powers: "
            "using separate 0 dBm FWM/SpRS bases plus cubic/linear power scaling"
        )
        set_power_override(0.0)
        with profile_scope("power +0 dBm: base FWM precompute_by_length"):
            fwm_base_by_idx, fwm_valid_ch = precompute_by_length(
                "fwm", specs, LENGTHS_KM, base_config,
                noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
            )
        with profile_scope("power +0 dBm: base SpRS precompute_by_length"):
            sprs_base_by_idx, sprs_valid_ch = precompute_by_length(
                "sprs", specs, LENGTHS_KM, base_config,
                noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
            )
        base_valid_ch = sorted(set(fwm_valid_ch) & set(sprs_valid_ch))
        results: dict[float, tuple[dict, list]] = {}
        for p in PRECOMPUTE_POWER_LEVELS:
            fwm_scale = _power_scaling_factor("fwm", float(p))
            sprs_scale = _power_scaling_factor("sprs", float(p))
            assert fwm_scale is not None and sprs_scale is not None
            with profile_scope(f"power {p:+.0f} dBm: combine FWM/SpRS bases"):
                all_by_idx = _combine_precomputed_results(
                    fwm_base_by_idx, sprs_base_by_idx, fwm_scale, sprs_scale
                )
            with profile_scope(f"power {p:+.0f} dBm: save CSV cache"):
                _cache_precomputed_result(all_by_idx, p, noise_type, list(specs.keys()), index_prefix="len")
            _POWER_CACHE[("len", float(p))] = all_by_idx
            results[float(p)] = (all_by_idx, base_valid_ch)

        set_power_override(0.0)
        return results[0.0]

    # Sequential fast path: avoids spawn overhead for fast noise types (e.g. sprs ~1s total)
    if not _MP_ENABLED or len(PRECOMPUTE_POWER_LEVELS) <= 2:
        results: dict[float, tuple[dict, list]] = {}
        for p in PRECOMPUTE_POWER_LEVELS:
            set_power_override(float(p))
            all_by_idx, valid_ch = precompute_by_length(
                noise_type, specs, LENGTHS_KM, base_config,
                noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
            )
            _cache_precomputed_result(all_by_idx, p, noise_type, list(specs.keys()), index_prefix="len")
            _POWER_CACHE[("len", p)] = all_by_idx
            results[p] = (all_by_idx, valid_ch)
    else:
        # Parallel: all power levels are independent
        n_workers = min(len(PRECOMPUTE_POWER_LEVELS), _get_mp_workers())
        ctx = multiprocessing.get_context()
        with ProcessPoolExecutor(max_workers=n_workers, mp_context=ctx) as executor:
            futures = {
                executor.submit(
                    _precompute_length_worker,
                    p, noise_type, specs, LENGTHS_KM, base_config,
                    noise_f_grid, osa_csv_path, fiber_params, osa_center_freq_hz,
                ): p
                for p in PRECOMPUTE_POWER_LEVELS
            }
            results = {}
            for fut in futures:
                p, all_by_idx, valid_ch = fut.result()
                # CSV caching is done inside the worker via precompute_by_length
                _POWER_CACHE[("len", p)] = all_by_idx
                results[p] = (all_by_idx, valid_ch)

    # Extract 0 dBm result
    result_all, result_valid = results[0.0]

# Restore 0 dBm in memory cache and power override
    set_power_override(0.0)
    _POWER_CACHE[("len", 0.0)] = result_all
    return result_all, result_valid


# =============================================================================
# Excel / simulation report export
# =============================================================================

def _interp_discrete_to_freq_grid(
    discrete_x: np.ndarray,  # channel center freqs (Hz), shape (N_q,)
    discrete_vals: np.ndarray,  # channel powers (W), shape (N_q,)
    freq_grid: np.ndarray,  # full frequency grid (Hz), shape (N_f,)
) -> np.ndarray:
    """Interpolate discrete channel-power data to full frequency grid.

    Returns array of same shape as freq_grid, with power at channel center
    frequencies and 0 elsewhere (boxcar interpolation).
    """
    out = np.zeros(len(freq_grid), dtype=np.float64)
    for x, v in zip(discrete_x, discrete_vals):
        idx = np.argmin(np.abs(freq_grid - x))
        out[idx] = v
    return out


def export_noise_vs_frequency_xlsx(
    all_by_ch: dict,
    model_keys: list[str],
    noise_f_grid: np.ndarray,
    LENGTHS_KM: np.ndarray,
    output_path: Path,
) -> None:
    """Export noise vs frequency data to noise_vs_frequency.xlsx.

    Each sheet = one fiber length.  Continuous models share full-frequency-grid
    columns; discrete model is interpolated to that grid (0 elsewhere).
    Only 0 dBm power is exported.
    """
    try:
        import openpyxl
    except ImportError:
        print("[export] openpyxl not available, skipping xlsx export")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for li in range(len(LENGTHS_KM)):
        length_km = float(LENGTHS_KM[li])
        sweep = all_by_ch.get(li, {})
        if not sweep:
            continue
        ws = wb.create_sheet(title=f"L={length_km:.0f}km")

        # Build header row: frequency + (fwd/bwd pairs per model)
        headers = ["frequency_THz"]
        for mk in model_keys:
            headers.append(f"{mk}_fwd_W")
            headers.append(f"{mk}_bwd_W")
        ws.append(headers)

        freq_thz = noise_f_grid / 1e12
        n_f = len(freq_thz)

        for fi in range(n_f):
            row = [freq_thz[fi]]
            for mk in model_keys:
                entry = sweep.get(mk, {})
                fwd = entry.get("fwd", np.array([]))
                bwd = entry.get("bwd", np.array([]))
                # Continuous: index directly into fwd/bwd arrays
                # Discrete: use channel_center x_kind
                y_kind = entry.get("y_kind", "")
                x_kind = entry.get("x_kind", "")
                if len(fwd) == n_f:
                    row.append(float(fwd[fi]) if fi < len(fwd) else 0.0)
                    row.append(float(bwd[fi]) if fi < len(bwd) else 0.0)
                else:
                    # Discrete model: only valid at channel center
                    x_data = np.asarray(entry.get("x", []), dtype=np.float64)
                    if y_kind == "channel_power" and len(x_data) == len(fwd):
                        # find closest channel center
                        target_f = noise_f_grid[fi]
                        idx = np.argmin(np.abs(x_data - target_f))
                        if np.abs(x_data[idx] - target_f) < 1e9:  # within 1 GHz
                            row.append(float(fwd[idx]))
                            row.append(float(bwd[idx]))
                        else:
                            row.append(0.0)
                            row.append(0.0)
                    else:
                        row.append(0.0)
                        row.append(0.0)
            ws.append(row)

    wb.save(output_path)
    print(f"[export] wrote {output_path}")


def export_noise_vs_length_xlsx(
    all_by_len: dict,
    model_keys: list[str],
    LENGTHS_KM: np.ndarray,
    quantum_center_freqs: np.ndarray,
    output_path: Path,
) -> None:
    """Export noise vs length data to noise_vs_length.xlsx.

    Each sheet = one quantum channel.  Only 0 dBm power is exported.
    """
    try:
        import openpyxl
    except ImportError:
        print("[export] openpyxl not available, skipping xlsx export")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for qi, qfreq in enumerate(quantum_center_freqs):
        ws = wb.create_sheet(title=f"C{qi + WDM_PARAMS['start_channel']}")

        headers = ["length_km"] + [f"{mk}_fwd_W" for mk in model_keys] + [f"{mk}_bwd_W" for mk in model_keys]
        ws.append(headers)

        for li, length_km in enumerate(LENGTHS_KM):
            row = [float(length_km)]
            for mk in model_keys:
                entry = all_by_len.get(qi, {}).get(mk, {})
                fwd = np.asarray(entry.get("fwd", []), dtype=np.float64)
                bwd = np.asarray(entry.get("bwd", []), dtype=np.float64)
                row.append(float(fwd[li]) if li < len(fwd) else 0.0)
                row.append(float(bwd[li]) if li < len(bwd) else 0.0)
            ws.append(row)

    wb.save(output_path)
    print(f"[export] wrote {output_path}")


def _to_dbm_scalar(value_w: float) -> float | None:
    """Convert linear power in watts to dBm, returning None for non-positive values."""
    if not np.isfinite(value_w) or value_w <= 0.0:
        return None
    return float(10.0 * np.log10(value_w / 1e-3))


def _normalize_itu_channel(channel_itu: float) -> int | float:
    """Normalize an ITU channel number to an int or half-integer when possible."""
    nearest_int = round(channel_itu)
    if np.isclose(channel_itu, nearest_int, atol=1e-9):
        return int(nearest_int)

    nearest_half = round(channel_itu * 2.0) / 2.0
    if np.isclose(channel_itu, nearest_half, atol=1e-9):
        return float(nearest_half)

    return float(channel_itu)


def _channel_label(channel_itu: int | float | None) -> str:
    """Return a display label like C39 or C32.5 for an ITU channel number."""
    if channel_itu is None:
        return ""
    if isinstance(channel_itu, float) and not channel_itu.is_integer():
        return f"C{channel_itu:.1f}"
    return f"C{int(channel_itu)}"


def _frequency_to_itu_channel(freq_hz: float) -> int | float:
    """Convert a center frequency to a 1-based ITU G.694.1 channel number."""
    channel_itu = (
        WDM_PARAMS["start_channel"]
        + (freq_hz - WDM_PARAMS["start_freq"]) / WDM_PARAMS["channel_spacing"]
    )
    return _normalize_itu_channel(float(channel_itu))


def _series_prefixes(model_keys: list[str], noise_type: str) -> list[tuple[str, str]]:
    """Return (model_key, direction) pairs used by exported CSV columns."""
    directions = ("fwd",) if noise_type == "only_signal" else ("fwd", "bwd")
    return [(model_key, direction) for model_key in model_keys for direction in directions]


def _metric_columns(prefix: str) -> list[str]:
    """Return exported metric columns for one model-direction series."""
    return [
        f"{prefix}_noise_power_w",
        f"{prefix}_noise_power_dbm",
        f"{prefix}_qber",
        f"{prefix}_skr_bit_per_pulse",
        f"{prefix}_skr_bps",
    ]


def _write_analysis_csv(
    filepath: Path,
    fieldnames: list[str],
    rows: list[dict[str, object]],
) -> None:
    """Write a user-facing CSV export with a header row."""
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"[export] wrote {filepath}")


def export_noise_vs_frequency_csv(
    all_by_ch_by_power: dict[float, dict],
    model_keys: list[str],
    specs: dict[str, dict],
    LENGTHS_KM: np.ndarray,
    noise_type: str,
    modulation_format: str,
    output_dir: Path,
    fiber_cfg=None,
    skr_cfg=None,
) -> dict[str, Path]:
    """Export noise-vs-frequency data as tidy and Origin-friendly CSV files.

    The tidy CSV stores one row per data point. The wide CSV stores one row per
    (power, fiber length, frequency) tuple and expands each model/direction into
    its own columns for direct plotting in Origin.

    If fiber_cfg and skr_cfg are provided, SKR/QBER columns are filled using
    the approximate finite key rate model.
    """
    modulation_slug = modulation_format.lower()
    base_name = f"noise_vs_frequency_{noise_type}_{modulation_slug}"
    tidy_path = output_dir / f"{base_name}_tidy.csv"
    wide_path = output_dir / f"{base_name}_wide.csv"
    series_pairs = _series_prefixes(model_keys, noise_type)

    tidy_rows: list[dict[str, object]] = []
    wide_rows: list[dict[str, object]] = []

    for power_dbm in sorted(all_by_ch_by_power):
        all_by_ch = all_by_ch_by_power[power_dbm]
        for l_idx in sorted(all_by_ch):
            sweep = all_by_ch[l_idx]
            length_km = float(LENGTHS_KM[l_idx])
            series_maps: dict[tuple[str, str], dict[int, float]] = {}
            x_lookup_hz: dict[int, float] = {}

            for model_key, direction in series_pairs:
                entry = sweep.get(model_key, {})
                x_data = np.asarray(entry.get("x", []), dtype=np.float64)
                y_data = np.asarray(entry.get(direction, []), dtype=np.float64)
                if len(x_data) == 0 or len(x_data) != len(y_data):
                    continue

                model_label = str(specs.get(model_key, {}).get("label", model_key))
                x_kind = str(entry.get("x_kind", ""))
                y_kind = str(entry.get("y_kind", ""))
                data_map: dict[int, float] = {}

                for x_hz, noise_power_w in zip(x_data, y_data):
                    x_key = int(round(float(x_hz)))
                    x_lookup_hz[x_key] = float(x_hz)
                    data_map[x_key] = float(noise_power_w)

                    channel_itu: int | float | str = ""
                    channel_label = ""
                    if x_kind == "channel_center":
                        channel_itu = _frequency_to_itu_channel(float(x_hz))
                        channel_label = _channel_label(channel_itu)

                    # Compute SKR/QBER if configs provided
                    _qber_val = ""
                    _skr_bpp_val = ""
                    _skr_bps_val = ""
                    if fiber_cfg is not None and skr_cfg is not None:
                        dist_m = float(LENGTHS_KM[l_idx] * 1000.0)
                        skr_d = compute_skr_point(float(noise_power_w), dist_m, float(x_hz), fiber_cfg, skr_cfg)
                        _, (bps, bpp, qber) = list(skr_d.items())[0] if skr_d else ("", (0.0, 0.0, float("nan")))
                        # Use approx_finite model as default
                        approx = skr_d.get("approx_finite", (0.0, 0.0, float("nan")))
                        _skr_bps_val = approx[0] if approx[0] > 0 else ""
                        _skr_bpp_val = approx[1] if approx[1] > 0 else ""
                        _qber_val = approx[2] if not np.isnan(float(approx[2])) else ""

                    tidy_rows.append(
                        {
                            "sweep_type": "vs_frequency",
                            "noise_type": noise_type,
                            "modulation_format": modulation_slug,
                            "power_dbm": float(power_dbm),
                            "fiber_length_km": length_km,
                            "model_key": model_key,
                            "model_label": model_label,
                            "direction": direction,
                            "x_kind": x_kind,
                            "y_kind": y_kind,
                            "frequency_thz": float(x_hz) / 1e12,
                            "channel_itu": channel_itu,
                            "channel_label": channel_label,
                            "noise_power_w": float(noise_power_w),
                            "noise_power_dbm": _to_dbm_scalar(float(noise_power_w)) or "",
                            "qber": _qber_val,
                            "skr_bit_per_pulse": _skr_bpp_val,
                            "skr_bps": _skr_bps_val,
                        }
                    )

                if data_map:
                    series_maps[(model_key, direction)] = data_map

            for x_key in sorted(x_lookup_hz):
                freq_hz = x_lookup_hz[x_key]
                channel_itu_value = _frequency_to_itu_channel(freq_hz)
                channel_itu: int | float | str = ""
                channel_label = ""
                if isinstance(channel_itu_value, int) or (
                    isinstance(channel_itu_value, float) and channel_itu_value.is_integer()
                ):
                    channel_itu = channel_itu_value
                    channel_label = _channel_label(channel_itu_value)

                row: dict[str, object] = {
                    "noise_type": noise_type,
                    "modulation_format": modulation_slug,
                    "power_dbm": float(power_dbm),
                    "fiber_length_km": length_km,
                    "frequency_thz": freq_hz / 1e12,
                    "channel_itu": channel_itu,
                    "channel_label": channel_label,
                }
                for model_key, direction in series_pairs:
                    prefix = f"{model_key}_{direction}"
                    noise_power_w = series_maps.get((model_key, direction), {}).get(x_key)
                    row[f"{prefix}_noise_power_w"] = (
                        float(noise_power_w) if noise_power_w is not None else ""
                    )
                    row[f"{prefix}_noise_power_dbm"] = (
                        _to_dbm_scalar(float(noise_power_w))
                        if noise_power_w is not None
                        else ""
                    ) or ""
                    # Compute SKR/QBER if configs provided
                    if fiber_cfg is not None and skr_cfg is not None and noise_power_w is not None:
                        dist_m = float(LENGTHS_KM[l_idx] * 1000.0)
                        skr_d = compute_skr_point(float(noise_power_w), dist_m, float(freq_hz), fiber_cfg, skr_cfg)
                        approx = skr_d.get("approx_finite", (0.0, 0.0, float("nan")))
                        row[f"{prefix}_qber"] = approx[2] if not np.isnan(float(approx[2])) else ""
                        row[f"{prefix}_skr_bit_per_pulse"] = approx[1] if approx[1] > 0 else ""
                        row[f"{prefix}_skr_bps"] = approx[0] if approx[0] > 0 else ""
                    else:
                        row[f"{prefix}_qber"] = ""
                        row[f"{prefix}_skr_bit_per_pulse"] = ""
                        row[f"{prefix}_skr_bps"] = ""
                wide_rows.append(row)

    tidy_fields = [
        "sweep_type",
        "noise_type",
        "modulation_format",
        "power_dbm",
        "fiber_length_km",
        "model_key",
        "model_label",
        "direction",
        "x_kind",
        "y_kind",
        "frequency_thz",
        "channel_itu",
        "channel_label",
        "noise_power_w",
        "noise_power_dbm",
        "qber",
        "skr_bit_per_pulse",
        "skr_bps",
    ]
    wide_fields = [
        "noise_type",
        "modulation_format",
        "power_dbm",
        "fiber_length_km",
        "frequency_thz",
        "channel_itu",
        "channel_label",
    ]
    for model_key, direction in series_pairs:
        wide_fields.extend(_metric_columns(f"{model_key}_{direction}"))

    _write_analysis_csv(tidy_path, tidy_fields, tidy_rows)
    _write_analysis_csv(wide_path, wide_fields, wide_rows)
    return {"tidy": tidy_path, "wide": wide_path}


def export_noise_vs_length_csv(
    all_by_len_by_power: dict[float, dict],
    model_keys: list[str],
    specs: dict[str, dict],
    LENGTHS_KM: np.ndarray,
    channel_index_lookup: dict[int, int],
    noise_type: str,
    modulation_format: str,
    output_dir: Path,
    fiber_cfg=None,
    skr_cfg=None,
) -> dict[str, Path]:
    """Export noise-vs-length data as tidy and Origin-friendly CSV files.

    If fiber_cfg and skr_cfg are provided, SKR/QBER columns are filled using
    the approximate finite key rate model.
    """
    modulation_slug = modulation_format.lower()
    base_name = f"noise_vs_length_{noise_type}_{modulation_slug}"
    tidy_path = output_dir / f"{base_name}_tidy.csv"
    wide_path = output_dir / f"{base_name}_wide.csv"
    series_pairs = _series_prefixes(model_keys, noise_type)

    tidy_rows: list[dict[str, object]] = []
    wide_rows: list[dict[str, object]] = []

    for power_dbm in sorted(all_by_len_by_power):
        all_by_len = all_by_len_by_power[power_dbm]
        for outer_idx in sorted(all_by_len):
            sweep = all_by_len[outer_idx]
            channel_itu = int(channel_index_lookup[outer_idx])
            channel_label = _channel_label(channel_itu)
            frequency_thz = (
                WDM_PARAMS["start_freq"]
                + (channel_itu - WDM_PARAMS["start_channel"]) * WDM_PARAMS["channel_spacing"]
            ) / 1e12

            for li, length_km in enumerate(LENGTHS_KM):
                row: dict[str, object] = {
                    "noise_type": noise_type,
                    "modulation_format": modulation_slug,
                    "power_dbm": float(power_dbm),
                    "channel_itu": channel_itu,
                    "channel_label": channel_label,
                    "frequency_thz": float(frequency_thz),
                    "fiber_length_km": float(length_km),
                }

                for model_key, direction in series_pairs:
                    entry = sweep.get(model_key, {})
                    y_data = np.asarray(entry.get(direction, []), dtype=np.float64)
                    noise_power_w = float(y_data[li]) if li < len(y_data) else None

                    # Compute SKR/QBER if configs provided
                    _qber_val = ""
                    _skr_bpp_val = ""
                    _skr_bps_val = ""
                    if fiber_cfg is not None and skr_cfg is not None and noise_power_w is not None:
                        dist_m = float(length_km * 1000.0)
                        f_q_hz = WDM_PARAMS["start_freq"] + (channel_itu - WDM_PARAMS["start_channel"]) * WDM_PARAMS["channel_spacing"]
                        skr_d = compute_skr_point(float(noise_power_w), dist_m, f_q_hz, fiber_cfg, skr_cfg)
                        approx = skr_d.get("approx_finite", (0.0, 0.0, float("nan")))
                        _skr_bps_val = approx[0] if approx[0] > 0 else ""
                        _skr_bpp_val = approx[1] if approx[1] > 0 else ""
                        _qber_val = approx[2] if not np.isnan(float(approx[2])) else ""

                    if noise_power_w is not None:
                        model_label = str(specs.get(model_key, {}).get("label", model_key))
                        tidy_rows.append(
                            {
                                "sweep_type": "vs_length",
                                "noise_type": noise_type,
                                "modulation_format": modulation_slug,
                                "power_dbm": float(power_dbm),
                                "channel_itu": channel_itu,
                                "channel_label": channel_label,
                                "frequency_thz": float(frequency_thz),
                                "fiber_length_km": float(length_km),
                                "model_key": model_key,
                                "model_label": model_label,
                                "direction": direction,
                                "noise_power_w": noise_power_w,
                                "noise_power_dbm": _to_dbm_scalar(noise_power_w) or "",
                                "qber": _qber_val,
                                "skr_bit_per_pulse": _skr_bpp_val,
                                "skr_bps": _skr_bps_val,
                            }
                        )

                    prefix = f"{model_key}_{direction}"
                    row[f"{prefix}_noise_power_w"] = noise_power_w if noise_power_w is not None else ""
                    row[f"{prefix}_noise_power_dbm"] = (
                        _to_dbm_scalar(noise_power_w) if noise_power_w is not None else ""
                    ) or ""
                    row[f"{prefix}_qber"] = _qber_val
                    row[f"{prefix}_skr_bit_per_pulse"] = _skr_bpp_val
                    row[f"{prefix}_skr_bps"] = _skr_bps_val

                wide_rows.append(row)

    tidy_fields = [
        "sweep_type",
        "noise_type",
        "modulation_format",
        "power_dbm",
        "channel_itu",
        "channel_label",
        "frequency_thz",
        "fiber_length_km",
        "model_key",
        "model_label",
        "direction",
        "noise_power_w",
        "noise_power_dbm",
        "qber",
        "skr_bit_per_pulse",
        "skr_bps",
    ]
    wide_fields = [
        "noise_type",
        "modulation_format",
        "power_dbm",
        "channel_itu",
        "channel_label",
        "frequency_thz",
        "fiber_length_km",
    ]
    for model_key, direction in series_pairs:
        wide_fields.extend(_metric_columns(f"{model_key}_{direction}"))

    _write_analysis_csv(tidy_path, tidy_fields, tidy_rows)
    _write_analysis_csv(wide_path, wide_fields, wide_rows)
    return {"tidy": tidy_path, "wide": wide_path}


def export_simulation_report(
    fiber_params: dict,
    noise_f_grid: np.ndarray,
    model_keys: list[str],
    noise_type: str,
    modulation_format: str,
    output_path: Path,
) -> None:
    """Write simulation parameters to simulation_report.txt."""
    lines = [
        "=== QKD Optical Network Simulation Report ===",
        f"Timestamp: 2026-04-17",
        f"Noise type: {noise_type}",
        f"Modulation format: {modulation_format}",
        "",
        "--- WDM Parameters ---",
        f"  start_freq: {WDM_PARAMS['start_freq']:.3e} Hz ({WDM_PARAMS['start_freq']/1e12:.1f} THz)",
        f"  end_channel: {WDM_PARAMS['end_channel']}",
        f"  channel_spacing: {WDM_PARAMS['channel_spacing']:.3e} Hz ({WDM_PARAMS['channel_spacing']/1e9:.0f} GHz)",
        f"  B_s: {WDM_PARAMS['B_s']:.3e} Hz ({WDM_PARAMS['B_s']/1e9:.1f} GBaud)",
        f"  P0: {WDM_PARAMS['P0']:.3e} W ({10*np.log10(WDM_PARAMS['P0']/1e-3):.1f} dBm)",
        f"  beta_rolloff: {WDM_PARAMS['beta_rolloff']}",
        f"  classical_channel_indices: {CLASSICAL_INDICES}",
        f"  classical_channel_ITU: {[f'C{ch}' for ch in CLASSICAL_INDICES]}",
        f"  classical_channel_freqs_THz: {[round(WDM_PARAMS['start_freq']/1e12 + (ch - WDM_PARAMS['start_channel']) * WDM_PARAMS['channel_spacing']/1e12, 4) for ch in CLASSICAL_INDICES]}",
        "",
        "--- Fiber Parameters ---",
        f"  alpha_dB_per_km: {fiber_params.get('alpha_dB_per_km', 'N/A')}",
        f"  gamma_per_W_km: {fiber_params.get('gamma_per_W_km', 'N/A')}",
        f"  D_ps_nm_km: {fiber_params.get('D_ps_nm_km', 'N/A')}",
        f"  D_slope_ps_nm2_km: {fiber_params.get('D_slope_ps_nm2_km', 'N/A')}",
        f"  A_eff: {fiber_params.get('A_eff', 'N/A')} m^2",
        f"  rayleigh_coeff: {fiber_params.get('rayleigh_coeff', 'N/A')}",
        f"  T_kelvin: {fiber_params.get('T_kelvin', 'N/A')} K",
        f"  length_km_samples: {list(fiber_params.get('length_km_samples', LENGTHS_KM))}",
        "",
        "--- Model Keys ---",
        f"  {model_keys}",
        "",
        "--- Frequency Grid ---",
        f"  noise_f_grid points: {len(noise_f_grid)}",
        f"  noise_f_grid range: {noise_f_grid[0]/1e12:.4f}  --  {noise_f_grid[-1]/1e12:.4f} THz",
        f"  NOISE_GRID_RESOLUTION_HZ: {NOISE_GRID_RESOLUTION_HZ:.3e}",
        f"  NOISE_FLOOR_W: {NOISE_FLOOR_W:.3e}",
    ]

    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[export] wrote {output_path}")
